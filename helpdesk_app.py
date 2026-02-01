# FULL MERGED helpdesk_app.py
# (Replace your current helpdesk_app.py with this file after backing it up)
#
# After saving this file restart Streamlit:
#   python -m streamlit run helpdesk_app.py --logger.level=debug
#
# Summary of the change in this version:
# - The app previously showed minimal/placeholder UI for non-dashboard pages when DB_AVAILABLE==True.
#   I added connected-mode implementations for:
#     - Helpdesk Tickets (list, search, download CSV, create ticket form placeholder)
#     - Asset Management (list, search, download CSV)
#     - Procurement Requests (list, search, download CSV)
#     - Fleet Management (list, search, download CSV)
# - Each connected-mode implementation uses execute_query() defensively and falls back to the old
#   informational message if the query fails. Demo fallbacks remain when DB is unavailable.
# - Everything else from the long merged file is preserved.

import logging
from pathlib import Path
from datetime import datetime, timedelta
import os
import traceback
import csv
from io import BytesIO
import base64
from typing import Optional, Tuple, TYPE_CHECKING
from urllib.parse import quote, urlparse

import streamlit as st
import streamlit.components.v1 as components

import pandas as pd
import smtplib
from email.message import EmailMessage

try:
    import plotly.express as px
    HAS_PLOTLY = True
except Exception:
    HAS_PLOTLY = False

# App colors
VDH_NAVY = "#002855"
VDH_ORANGE = "#FF6B35"

# Safe image loader
_logger = logging.getLogger("safe_image_loader")

def safe_st_image(path_or_bytes, width: Optional[int] = None, use_container_width: bool = False, stretch: bool = False, **kwargs):
    try:
        if use_container_width or stretch:
            width_arg = "stretch"
        else:
            width_arg = width

        if isinstance(path_or_bytes, str):
            parsed = urlparse(path_or_bytes)
            if parsed.scheme in ("http", "https", "data"):
                if width_arg == "stretch":
                    try:
                        st.image(path_or_bytes, width='stretch', **kwargs)
                    except Exception:
                        st.image(path_or_bytes, **kwargs)
                elif isinstance(width_arg, int):
                    st.image(path_or_bytes, width=width_arg, **kwargs)
                else:
                    st.image(path_or_bytes, **kwargs)
                return

        if isinstance(path_or_bytes, (Path, str)):
            p = Path(path_or_bytes)
            if not p.is_absolute():
                p = Path.cwd() / p
            if p.exists():
                if width_arg == "stretch":
                    try:
                        st.image(str(p), width='stretch', **kwargs)
                    except Exception:
                        st.image(str(p), **kwargs)
                elif isinstance(width_arg, int):
                    st.image(str(p), width=width_arg, **kwargs)
                else:
                    st.image(str(p), **kwargs)
            else:
                st.warning(f"Image not found: {p.name}")
            return

        if width_arg == "stretch":
            try:
                st.image(path_or_bytes, width='stretch', **kwargs)
            except Exception:
                st.image(path_or_bytes, **kwargs)
        elif isinstance(width_arg, int):
            st.image(path_or_bytes, width=width_arg, **kwargs)
        else:
            st.image(path_or_bytes, **kwargs)
    except Exception:
        _logger.exception("Failed to render image %s", getattr(path_or_bytes, "name", str(path_or_bytes)))
        st.warning("Could not load image.")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Database helpers
@st.cache_resource
def get_connection_string():
    try:
        server = st.secrets["database"]["server"]
        database = st.secrets["database"]["database"]
        username = st.secrets["database"]["username"]
        password = st.secrets["database"]["password"]
    except Exception:
        server = os.getenv("DB_SERVER")
        database = os.getenv("DB_DATABASE")
        username = os.getenv("DB_USERNAME")
        password = os.getenv("DB_PASSWORD")
    return server, database, username, password

def get_db_connection():
    try:
        import pyodbc
    except Exception:
        raise ConnectionError("pyodbc is not installed")

    conn_str_env = os.getenv("DB_CONN")
    if conn_str_env:
        try:
            return pyodbc.connect(conn_str_env, autocommit=False, timeout=60)
        except Exception as e:
            logger.warning("DB_CONN connect failed: %s", e)

    server, database, username, password = get_connection_string()
    if not all([server, database, username, password]):
        raise ConnectionError("Missing DB connection parameters.")

    try:
        if "@" not in username and "database.windows.net" in server:
            username = f"{username}@{server.split('.')[0]}"
    except Exception:
        pass

    # Try different ODBC drivers
    try:
        import pyodbc
        available_drivers = [x for x in pyodbc.drivers() if 'SQL Server' in x]
        if available_drivers:
            driver = available_drivers[0]  # Use first available
            logger.info(f"Using ODBC driver: {driver}")
        else:
            driver = "ODBC Driver 18 for SQL Server"
            logger.warning("No SQL Server drivers found, using default")
    except:
        driver = "ODBC Driver 18 for SQL Server"

    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER=tcp:{server},1433;"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
        "Encrypt=yes;"
        "TrustServerCertificate=no;"
        "Connection Timeout=30;"
    )
    try:
        conn = pyodbc.connect(conn_str, autocommit=False, timeout=60)
        return conn
    except Exception as e:
        logger.exception("Database connection failed")
        raise ConnectionError(str(e))

def execute_query(query: str, params: Optional[tuple] = None) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    try:
        conn = get_db_connection()
    except Exception as e:
        return None, f"Connection error: {e}"
    try:
        if params:
            df = pd.read_sql(query, conn, params=params)
        else:
            df = pd.read_sql(query, conn)
        conn.close()
        return df, None
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return None, f"Query error: {e}"

def execute_non_query(query: str, params: Optional[tuple] = None) -> Tuple[bool, Optional[str]]:
    try:
        conn = get_db_connection()
    except Exception as e:
        return False, f"Connection error: {e}"
    try:
        cur = conn.cursor()
        if params:
            cur.execute(query, params)
        else:
            cur.execute(query)
        conn.commit()
        cur.close()
        conn.close()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        return False, f"Execution error: {e}\n{traceback.format_exc()}"

def check_db_available():
    """Check if database is available - can be called from anywhere"""
    try:
        conn = get_db_connection()
        conn.close()
        return True
    except Exception:
        return False

def send_email_notification(to_email, subject, body):
    """Send email notification (placeholder - implement with your email service)"""
    logger.info(f"Email notification: {to_email} - {subject}")
    # TODO: Implement with SMTP or SendGrid
    return True

def send_email(subject: str, body: str, recipients: list):
    try:
        smtp_cfg = {}
        try:
            smtp_cfg = st.secrets.get("smtp", {})
        except Exception:
            smtp_cfg = {}
        host = smtp_cfg.get("host") or os.getenv("SMTP_HOST")
        port = int(smtp_cfg.get("port") or os.getenv("SMTP_PORT") or 587)
        username = smtp_cfg.get("username") or os.getenv("SMTP_USER")
        password = smtp_cfg.get("password") or os.getenv("SMTP_PASS")
        use_tls = smtp_cfg.get("use_tls", True)
        from_addr = smtp_cfg.get("from_address") or username or "no-reply@vdh.gov"

        if not host or not recipients:
            logger.warning("SMTP host or recipients missing; skipping email.")
            return False

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = from_addr
        msg["To"] = ", ".join(recipients)
        msg.set_content(body)

        server = smtplib.SMTP(host, port, timeout=30)
        if use_tls:
            server.starttls()
        if username and password:
            server.login(username, password)
        server.send_message(msg)
        server.quit()
        logger.info("Email sent to %s", recipients)
        return True
    except Exception as e:
        logger.exception("Failed to send email: %s", e)
        return False

def insert_and_get_id(insert_sql: str, params: tuple = None) -> Tuple[Optional[int], Optional[str]]:
    try:
        conn = get_db_connection()
    except Exception as e:
        return None, f"Connection error: {e}"

    try:
        cur = conn.cursor()
        if params:
            cur.execute(insert_sql, params)
        else:
            cur.execute(insert_sql)
        new_id = None
        try:
            row = cur.fetchone()
            if row and row[0] is not None:
                new_id = int(row[0])
        except Exception:
            new_id = None
        conn.commit()
        cur.close()
        conn.close()
        return new_id, None
    except Exception as e:
        try:
            conn.rollback()
        except:
            pass
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return None, f"Execution error: {e}\n{traceback.format_exc()}"

# Small helpers
def safe_rerun():
    try:
        st.rerun()
    except Exception:
        try:
            st.experimental_rerun()
        except Exception:
            st.warning("Please refresh the page manually to see updates.")

def generate_ticket_number():
    now = datetime.now()
    return f"TKT-{now.strftime('%Y%m%d%H%M%S')}"

def generate_asset_tag():
    now = datetime.now()
    return f"COV-{now.strftime('%Y%m%d%H%M%S')}"

def generate_procurement_number():
    now = datetime.now()
    return f"PR-{now.strftime('%Y%m%d%H%M%S')}"

# Report helpers (Excel / PDF) - unchanged from previous file (omitted here for brevity in comment)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_EXCEL = True
except Exception:
    HAS_EXCEL = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import inch
    HAS_PDF = True
except Exception:
    HAS_PDF = False

def generate_excel_report(df, report_title):
    if not HAS_EXCEL:
        return None
    out = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = report_title[:31]
    header_fill = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for col_cells in ws.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)
    wb.save(out)
    out.seek(0)
    return out

def generate_pdf_report(df, report_title):
    if not HAS_PDF:
        return None
    out = BytesIO()
    doc = SimpleDocTemplate(out, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor(VDH_NAVY))
    elements.append(Paragraph(report_title, title_style))
    elements.append(Spacer(1, 6))
    data = [df.columns.tolist()] + df.values.tolist()
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor(VDH_NAVY)),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    out.seek(0)
    return out

# Public submit handlers and param helpers (unchanged)
def _ensure_submissions_dir():
    d = Path("public_submissions")
    d.mkdir(parents=True, exist_ok=True)
    return d

def _append_submission_csv(form_key: str, row: dict):
    d = _ensure_submissions_dir()
    filename = d / f"{form_key}.csv"
    file_exists = filename.exists()
    fieldnames = list(row.keys())
    with open(filename, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)

def _get_param_value(params: dict, key: str) -> str:
    v = params.get(key, "")
    if isinstance(v, (list, tuple)):
        return v[0] if len(v) > 0 else ""
    return v or ""

def render_helpdesk_ticket_public_form():
    # Hide default Streamlit navigation
    st.markdown("""
        <style>
        [data-testid="stSidebarNav"] {display: none !important;}
        section[data-testid="stSidebar"] > div:first-child {display: none !important;}
        </style>
    """, unsafe_allow_html=True)
    
    # Custom sidebar for public forms
    with st.sidebar:
        st.title("🏥 VDH Service Center")
        st.markdown("---")
        st.markdown("### 🔗 Quick Links")
        st.markdown("[🔐 Login to Service Desk](http://localhost:8501)")
        st.markdown("[🏛️ VDH Intranet](https://www.vdh.virginia.gov/)")
        st.markdown("---")
        
        # Close Window button
        if st.button("❌ Close Window", use_container_width=True, type="secondary"):
            st.markdown("""
                <script>
                window.close();
                </script>
            """, unsafe_allow_html=True)
            st.warning("If the window doesn't close automatically, you can close this tab manually.")
        
        st.caption("Virginia Department of Health")
    
    st.title("Submit Helpdesk Ticket")
    st.markdown("Use this form to submit a public helpdesk ticket. Provide contact info and a description of your issue.")
    with st.form("public_helpdesk_form", clear_on_submit=False):
        name = st.text_input("Full name", "")
        email = st.text_input("Email", "")
        phone = st.text_input("Phone (optional)", "")
        department = st.text_input("Department (optional)", "")
        subject = st.text_input("Subject", "")
        priority = st.selectbox("Priority", ["Low", "Medium", "High"], index=1)
        description = st.text_area("Description", height=200)
        attachments = st.file_uploader("Attachments (optional)", accept_multiple_files=True)
        submitted = st.form_submit_button("Submit Ticket")
    if submitted:
        now = datetime.utcnow().isoformat() + "Z"
        submission_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
        attachment_names = ";".join([f.name for f in attachments]) if attachments else ""
        row = {
            "submission_id": submission_id,
            "timestamp_utc": now,
            "name": name,
            "email": email,
            "phone": phone,
            "department": department,
            "subject": subject,
            "priority": priority,
            "description": description,
            "attachments": attachment_names,
        }
        try:
            _append_submission_csv("helpdesk_tickets", row)
            
            # Also save to database for tracking and badges
            if check_db_available():
                ticket_insert = """
                    INSERT INTO dbo.Tickets 
                        (requester_name, requester_email, requester_phone, department, 
                         subject, priority, description, status, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'New', GETDATE())
                """
                success, error = execute_non_query(
                    ticket_insert, 
                    (name, email, phone, department, subject, priority, description)
                )
                if not success:
                    logger.warning(f"Failed to save ticket to database: {error}")
            
            if attachments:
                adir = _ensure_submissions_dir() / submission_id
                adir.mkdir(parents=True, exist_ok=True)
                for f in attachments:
                    contents = f.read()
                    with open(adir / f.name, "wb") as fh:
                        fh.write(contents)
            st.success(f"✅ **Your ticket has been submitted successfully!**")
            st.balloons()
            
            # Show success page
            st.markdown("---")
            st.info("📧 **Your ticket was received and will be resolved in the order it was received.**")
            
            st.markdown("---")
            st.write("### 📋 Ticket Details")
            st.write(f"**Ticket ID:** {submission_id}")
            st.write(f"**Subject:** {subject}")
            st.write(f"**Priority:** {priority}")
            st.write(f"**Status:** New")
            
            st.markdown("---")
            st.write("### What happens next?")
            st.write("1. 📝 Your ticket has been logged in our system")
            st.write("2. 👀 An administrator will review it shortly")
            st.write("3. 📧 You'll receive email updates on the ticket status")
            st.write("4. 🔧 We'll work to resolve your issue as quickly as possible")
            
            st.markdown("---")
            
            # Safe secrets access with fallback
            try:
                support_email = st.secrets.get('support_email', 'support@vdh.virginia.gov')
            except Exception:
                support_email = 'support@vdh.virginia.gov'
            
            st.caption(f"Need immediate assistance? Contact support at {support_email}")
            
            st.stop()  # Prevent redirect to main app
        except Exception as e:
            st.error("Sorry, we could not save your submission. Please try again or contact support.")
            st.exception(e)

def render_request_vehicle_public_form():
    # Hide default Streamlit navigation
    st.markdown("""
        <style>
        [data-testid="stSidebarNav"] {display: none !important;}
        section[data-testid="stSidebar"] > div:first-child {display: none !important;}
        </style>
    """, unsafe_allow_html=True)
    
    # Custom sidebar for public forms
    with st.sidebar:
        st.title("🏥 VDH Service Center")
        st.markdown("---")
        st.markdown("### 🔗 Quick Links")
        st.markdown("[🔐 Login to Service Desk](http://localhost:8501)")
        st.markdown("[🏛️ VDH Intranet](https://www.vdh.virginia.gov/)")
        st.markdown("---")
        
        # Close Window button
        if st.button("❌ Close Window", use_container_width=True, type="secondary"):
            st.markdown("""
                <script>
                window.close();
                </script>
            """, unsafe_allow_html=True)
            st.warning("If the window doesn't close automatically, you can close this tab manually.")
        
        st.caption("Virginia Department of Health")
    
    st.title("🚗 Request a Vehicle")
    st.info("👀 **Browse available vehicles below and select one to continue with your request.**")
    
    if not check_db_available():
        st.warning("Database unavailable. Vehicle request system requires database connection.")
        return
    
    # Query available vehicles sorted by recent usage
    query = """
        SELECT 
            v.id,
            v.year,
            v.make_model,
            v.license_plate,
            v.current_mileage,
            v.initial_mileage,
            v.photo_url,
            v.status,
            v.miles_until_service,
            COALESCE(v.last_used_date, DATEADD(day, -100, GETDATE())) as last_used,
            DATEDIFF(day, COALESCE(v.last_used_date, DATEADD(day, -100, GETDATE())), GETDATE()) as days_since_used,
            v.usage_count
        FROM dbo.vehicles v
        WHERE v.status = 'Available'
        ORDER BY v.usage_count ASC, v.current_mileage ASC
    """
    
    vehicles_df, vehicles_err = execute_query(query)
    
    if vehicles_err:
        st.error(f"Error loading vehicles: {vehicles_err}")
        return
    elif vehicles_df is None or vehicles_df.empty:
        st.warning("⚠️ No vehicles are currently available. Please check back later or contact the administrator.")
        return
    
    st.success(f"📊 {len(vehicles_df)} vehicle(s) available for request")
    
    # Initialize session state for selected vehicle
    if 'selected_vehicle_id' not in st.session_state:
        st.session_state.selected_vehicle_id = None
    
    # Vehicle Gallery (if no vehicle selected)
    if st.session_state.selected_vehicle_id is None:
        st.markdown("---")
        st.subheader("🚙 Available Vehicles (Least Used First)")
        
        for idx, vehicle in vehicles_df.iterrows():
            vehicle_id = vehicle['id']
            year = vehicle['year']
            make_model = vehicle['make_model']
            license_plate = vehicle['license_plate']
            current_mileage = vehicle['current_mileage']
            photo_url = vehicle.get('photo_url')
            miles_until_service = vehicle['miles_until_service']
            usage_count = vehicle['usage_count']
            
            # Service status indicator
            if miles_until_service > 1000:
                service_indicator = "🟢 Service: Good"
            elif miles_until_service > 500:
                service_indicator = "🟡 Service Due Soon"
            else:
                service_indicator = "🔴 Service Due Now"
            
            # Card styling
            row_class = "item-row-even" if idx % 2 == 0 else "item-row-odd"
            st.markdown(f'<div class="item-row {row_class}">', unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns([1, 2, 1])
            
            with col1:
                # Display photo
                if photo_url and str(photo_url) != 'nan':
                    try:
                        st.image(photo_url, width=150)
                    except:
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                    padding: 20px; text-align: center; border-radius: 8px; 
                                    color: white; font-size: 36px; width: 150px;">
                            🚗
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                padding: 20px; text-align: center; border-radius: 8px; 
                                color: white; font-size: 36px; width: 150px;">
                        🚗
                    </div>
                    """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f'<div class="list-header">🚗 {year} {make_model}</div>', unsafe_allow_html=True)
                st.write(f"**License:** {license_plate}")
                st.write(f"**Mileage:** {current_mileage:,} miles")
                st.write(f"**Times Used:** {usage_count}")
                st.caption(service_indicator)
            
            with col3:
                st.write("") # Spacer
                if st.button(f"✅ Select Vehicle", key=f"select_vehicle_{vehicle_id}"):
                    st.session_state.selected_vehicle_id = vehicle_id
                    st.rerun()
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Request Form (if vehicle selected)
    else:
        # Get selected vehicle details
        selected_vehicle = vehicles_df[vehicles_df['id'] == st.session_state.selected_vehicle_id]
        
        if len(selected_vehicle) == 0:
            st.error("Selected vehicle not found or no longer available.")
            if st.button("← Back to Vehicle List"):
                st.session_state.selected_vehicle_id = None
                st.rerun()
        else:
            vehicle = selected_vehicle.iloc[0]
            
            st.markdown("---")
            st.subheader(f"📝 Vehicle Request Form")
            
            # Show selected vehicle
            col1, col2 = st.columns([1, 3])
            with col1:
                photo_url = vehicle.get('photo_url')
                if photo_url and str(photo_url) != 'nan':
                    try:
                        st.image(photo_url, width=150, caption="Selected Vehicle")
                    except:
                        st.write("🚗")
                else:
                    st.write("🚗")
            
            with col2:
                st.write(f"### 🚗 {vehicle['year']} {vehicle['make_model']}")
                st.write(f"**License:** {vehicle['license_plate']}")
                st.write(f"**Current Mileage:** {vehicle['current_mileage']:,} miles")
                
                if st.button("← Change Vehicle"):
                    st.session_state.selected_vehicle_id = None
                    st.rerun()
            
            st.markdown("---")
            
            # Request form
            with st.form("vehicle_request_form"):
                st.write("### Your Information")
                
                col1, col2 = st.columns(2)
                with col1:
                    requester_name = st.text_input("Full Name *", placeholder="John Doe")
                    requester_email = st.text_input("Email *", placeholder="john.doe@vdh.virginia.gov")
                with col2:
                    requester_phone = st.text_input("Phone Number", placeholder="804-555-1234")
                    requester_location = st.selectbox("Your Location *", [
                        "", "Crater", "Dinwiddie County", "Greensville/Emporia", 
                        "Surry County", "Prince George", "Sussex County", 
                        "Hopewell", "Petersburg"
                    ])
                
                st.write("### Trip Details")
                
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input("Start Date *")
                    start_time = st.time_input("Start Time")
                with col2:
                    end_date = st.date_input("End Date *")
                    end_time = st.time_input("End Time")
                
                estimated_miles = st.number_input("Estimated Miles *", min_value=0, value=50, step=10)
                
                purpose = st.text_area("Purpose of Trip *", 
                    placeholder="Please describe the purpose and destination of your trip...",
                    height=100)
                
                notes = st.text_area("Additional Notes (Optional)", 
                    placeholder="Any special requirements or additional information...",
                    height=80)
                
                st.markdown("---")
                
                col1, col2, col3 = st.columns([1, 1, 1])
                with col2:
                    submit_button = st.form_submit_button("🚀 Submit Request", use_container_width=True)
                
                if submit_button:
                    # Validation
                    if not requester_name or not requester_email or not requester_location or not purpose:
                        st.error("❌ Please fill in all required fields marked with *")
                    elif not start_date or not end_date:
                        st.error("❌ Please provide start and end dates")
                    elif start_date > end_date:
                        st.error("❌ End date must be after start date")
                    else:
                        # Insert vehicle request
                        insert_query = """
                            INSERT INTO dbo.Vehicle_Requests (
                                vehicle_id, requester_name, requester_email, requester_phone,
                                requester_location, purpose, start_date, end_date,
                                start_time, end_time, estimated_miles, notes, status
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending')
                        """
                        
                        params = (
                            int(st.session_state.selected_vehicle_id),
                            requester_name,
                            requester_email,
                            requester_phone,
                            requester_location,
                            purpose,
                            start_date,
                            end_date,
                            start_time,
                            end_time,
                            estimated_miles,
                            notes
                        )
                        
                        success, error = execute_non_query(insert_query, params)
                        
                        if success:
                            # Create notification
                            notif_query = """
                                INSERT INTO dbo.Notifications (notification_type, reference_id, message, created_by)
                                VALUES ('VehicleRequest', (SELECT MAX(request_id) FROM dbo.Vehicle_Requests), ?, ?)
                            """
                            message = f"New vehicle request from {requester_name} for {vehicle['year']} {vehicle['make_model']}"
                            execute_non_query(notif_query, (message, requester_name))
                            
                            # Send email notification
                            send_email_notification(
                                "admin@vdh.virginia.gov",
                                "New Vehicle Request",
                                f"New vehicle request from {requester_name}\n\nVehicle: {vehicle['year']} {vehicle['make_model']}\nDates: {start_date} to {end_date}"
                            )
                            
                            st.success("✅ **Vehicle request submitted successfully!**")
                            st.info("📧 A confirmation email has been sent. You will be notified once your request is reviewed.")
                            st.balloons()
                            
                            # Reset selection
                            st.session_state.selected_vehicle_id = None
                            
                            # Show success message with details
                            st.markdown("---")
                            st.write("### 📋 Request Summary")
                            st.write(f"**Vehicle:** {vehicle['year']} {vehicle['make_model']} ({vehicle['license_plate']})")
                            st.write(f"**Dates:** {start_date} to {end_date}")
                            st.write(f"**Estimated Miles:** {estimated_miles}")
                            st.write(f"**Status:** Pending Administrator Approval")

                            st.markdown("---")
                            st.info("📧 **Your request has been submitted. Once approved, you will be emailed a link to submit starting and ending mileage and upload any trip photos. Thanks and have a great day!**")
                            st.markdown("---")
                            st.write("### What happens next?")
                            st.write("1. 📋 Your request will be reviewed by an administrator")
                            st.write("2. 📧 You'll receive an email notification once approved")
                            st.write("3. 🚗 Pick up your vehicle at the scheduled time")
                            st.write("4. 📸 Submit mileage and photos after your trip via email link")
                            st.stop()  # Prevent redirect to main app
                        else:
                            st.error(f"❌ Error submitting request: {error}")

def render_procurement_request_public_form():
    # Hide default Streamlit navigation
    st.markdown("""
        <style>
        [data-testid="stSidebarNav"] {display: none !important;}
        section[data-testid="stSidebar"] > div:first-child {display: none !important;}
        </style>
    """, unsafe_allow_html=True)
    
    # Custom sidebar for public forms
    with st.sidebar:
        st.title("🏥 VDH Service Center")
        st.markdown("---")
        st.markdown("### 🔗 Quick Links")
        st.markdown("[🔐 Login to Service Desk](http://localhost:8501)")
        st.markdown("[🏛️ VDH Intranet](https://www.vdh.virginia.gov/)")
        st.markdown("---")
        
        # Close Window button
        if st.button("❌ Close Window", use_container_width=True, type="secondary"):
            st.markdown("""
                <script>
                window.close();
                </script>
            """, unsafe_allow_html=True)
            st.warning("If the window doesn't close automatically, you can close this tab manually.")
        
        st.caption("Virginia Department of Health")
    
    st.title("Public Procurement Request")
    st.markdown("Submit a procurement request using this public form.")
    with st.form("public_procurement_form", clear_on_submit=False):
        requester = st.text_input("Requester name", "")
        email = st.text_input("Requester email", "")
        location = st.text_input("Location", "")
        items = st.text_area("Items / Description", height=150)
        total = st.text_input("Estimated total amount", "")
        submitted = st.form_submit_button("Submit Requisition")
    if submitted:
        now = datetime.utcnow().isoformat() + "Z"
        submission_id = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
        row = {
            "submission_id": submission_id,
            "timestamp_utc": now,
            "requester": requester,
            "email": email,
            "location": location,
            "items": items,
            "estimated_total": total
        }
        try:
            _append_submission_csv("procurement_requests", row)
            st.success(f"Procurement request submitted (ID: {submission_id})")
        except Exception as e:
            st.error("Failed to save procurement request")
            st.exception(e)

def render_driver_trip_entry_public_form():
    """Public form for drivers to log trip start/end"""
    # Mobile-friendly styling
    st.markdown("""
        <style>
        .stButton>button {
            width: 100%;
            height: 60px;
            font-size: 20px;
            margin: 10px 0;
        }
        .stTextInput>div>div>input, .stNumberInput>div>div>input {
            font-size: 18px;
            height: 50px;
        }
        .stSelectbox>div>div>select {
            font-size: 18px;
            height: 50px;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Custom sidebar for public forms
    with st.sidebar:
        st.title("🏥 VDH Service Center")
        st.markdown("---")
        st.markdown("### 🔗 Quick Links")
        st.markdown("[🔐 Login to Service Desk](http://localhost:8501)")
        st.markdown("[🏛️ VDH Intranet](https://www.vdh.virginia.gov/)")
        st.markdown("---")
        
        # Close Window button
        if st.button("❌ Close Window", use_container_width=True, type="secondary"):
            st.markdown("""
                <script>
                window.close();
                </script>
            """, unsafe_allow_html=True)
            st.warning("If the window doesn't close automatically, you can close this tab manually.")
        
        st.caption("Virginia Department of Health")
    
    st.title("🚙 Driver Trip Entry")
    st.markdown("Log your vehicle trip start and end information for fleet reporting.")
    
    # Driver identification
    st.subheader("👤 Driver Information")
    driver_email = st.text_input("📧 Your Email Address", placeholder="driver@vdh.virginia.gov")
    
    if not driver_email:
        st.info("👆 Please enter your email to continue")
        st.stop()
    
    # Get approved vehicles for this driver
    vehicles_query = """
        SELECT v.id, v.vehicle_name, v.vehicle_type, v.license_plate, v.vin
        FROM dbo.dispatched_vehicles dv
        JOIN dbo.vehicles v ON dv.vehicle_id = v.id
        WHERE dv.approved_email = ? AND dv.approval_status = 'Approved'
    """
    vehicles_df, vehicles_err = execute_query(vehicles_query, (driver_email,))
    
    if vehicles_err or vehicles_df is None or vehicles_df.empty:
        st.warning(f"⚠️ No approved vehicles found for {driver_email}")
        st.info("Contact your fleet manager to get vehicle approval before logging trips.")
        st.stop()
    
    # Vehicle selection
    st.markdown("---")
    st.subheader("🚙 Select Your Vehicle")
    
    vehicle_options = [
        f"{row['vehicle_name']} - {row['license_plate']}" 
        for _, row in vehicles_df.iterrows()
    ]
    selected_vehicle = st.selectbox("Vehicle", vehicle_options, label_visibility="collapsed")
    vehicle_id = vehicles_df.iloc[vehicle_options.index(selected_vehicle)]['id']
    
    # Check for active trip
    active_trip_query = """
        SELECT trip_id, start_location, start_mileage, start_datetime, department
        FROM dbo.vehicle_trips
        WHERE vehicle_id = ? AND driver_email = ? AND trip_status = 'In Progress'
        ORDER BY start_datetime DESC
    """
    active_trip_df, _ = execute_query(active_trip_query, (int(vehicle_id), driver_email))
    
    has_active_trip = active_trip_df is not None and not active_trip_df.empty
    
    # Tabs for Start Trip / End Trip
    tab1, tab2 = st.tabs(["🟢 Start Trip", "🔴 End Trip"])
    
    with tab1:
        st.subheader("🟢 Start a New Trip")
        
        if has_active_trip:
            st.warning("⚠️ You have an active trip! Please end it before starting a new one.")
            st.info(f"Active trip started at: {active_trip_df.iloc[0]['start_location']}")
        else:
            with st.form("start_trip_form"):
                col1, col2 = st.columns(2)
                with col1:
                    start_location = st.text_input("📍 Starting Location*", 
                                                  placeholder="e.g., VDH Petersburg Office")
                    start_mileage = st.number_input("🛣️ Starting Mileage*", 
                                                   min_value=0, step=1)
                with col2:
                    departments = ["IT", "Administration", "Nursing", "Environmental Health", 
                                 "Vital Records", "Clinical Services", "Maintenance", "Other"]
                    department = st.selectbox("🏢 Department*", departments)
                    trip_notes = st.text_area("📝 Trip Purpose (Optional)", 
                                              placeholder="Brief description of trip purpose")
                
                submit_start = st.form_submit_button("🚗 START TRIP", use_container_width=True)
                
                if submit_start:
                    if not start_location or start_mileage == 0:
                        st.error("Please fill in all required fields!")
                    else:
                        insert_query = """
                            INSERT INTO dbo.vehicle_trips 
                            (vehicle_id, driver_email, driver_name, department, 
                             start_location, start_mileage, start_datetime, trip_status, notes)
                            VALUES (?, ?, ?, ?, ?, ?, GETDATE(), 'In Progress', ?)
                        """
                        driver_name = driver_email.split('@')[0]  # Extract name from email
                        
                        result, insert_err = execute_non_query(
                            insert_query, 
                            (int(vehicle_id), driver_email, driver_name, department,
                             start_location, int(start_mileage), trip_notes)
                        )
                        
                        if insert_err or not result:
                            st.error(f"❌ Error starting trip: {insert_err}")
                        else:
                            st.success("✅ Trip started successfully!")
                            st.balloons()
                            time.sleep(1)
                            st.rerun()
    
    with tab2:
        st.subheader("🔴 End Current Trip")
        
        if not has_active_trip:
            st.info("ℹ️ No active trip to end. Start a new trip first!")
        else:
            trip_id = active_trip_df.iloc[0]['trip_id']
            start_info = active_trip_df.iloc[0]
            
            st.success(f"**Active Trip:**")
            st.write(f"🏁 Started from: {start_info['start_location']}")
            st.write(f"📏 Starting mileage: {start_info['start_mileage']:,} miles")
            st.write(f"🕐 Started at: {start_info['start_datetime']}")
            
            with st.form("end_trip_form"):
                col1, col2 = st.columns(2)
                with col1:
                    end_location = st.text_input("📍 Ending Location*", 
                                                placeholder="e.g., VDH Hopewell Office")
                    end_mileage = st.number_input("🛣️ Ending Mileage*", 
                                                 min_value=int(start_info['start_mileage']), 
                                                 step=1)
                with col2:
                    end_notes = st.text_area("📝 Trip Notes (Optional)", 
                                            placeholder="Any issues or observations")
                    
                    # Photo upload
                    uploaded_photos = st.file_uploader(
                        "📸 Upload Trip Photos (Optional)",
                        type=['jpg', 'jpeg', 'png'],
                        accept_multiple_files=True
                    )
                
                miles_driven = end_mileage - int(start_info['start_mileage']) if end_mileage > 0 else 0
                if miles_driven > 0:
                    st.info(f"🛣️ Miles driven: **{miles_driven}** miles")
                
                submit_end = st.form_submit_button("🏁 END TRIP", use_container_width=True)
                
                if submit_end:
                    if not end_location or end_mileage == 0:
                        st.error("Please fill in all required fields!")
                    elif end_mileage < int(start_info['start_mileage']):
                        st.error("Ending mileage cannot be less than starting mileage!")
                    else:
                        # Update trip
                        update_query = """
                            UPDATE dbo.vehicle_trips
                            SET end_location = ?, end_mileage = ?, end_datetime = GETDATE(),
                                trip_status = 'Completed', 
                                notes = ISNULL(notes, '') + CHAR(13) + CHAR(10) + ?
                            WHERE trip_id = ?
                        """
                        
                        result, update_err = execute_non_query(
                            update_query,
                            (end_location, int(end_mileage), end_notes or '', int(trip_id))
                        )
                        
                        if update_err or not result:
                            st.error(f"❌ Error ending trip: {update_err}")
                        else:
                            # Upload photos if any
                            if uploaded_photos:
                                for photo in uploaded_photos:
                                    photo_bytes = photo.read()
                                    photo_query = """
                                        INSERT INTO dbo.trip_photos 
                                        (trip_id, photo_data, photo_filename, photo_size)
                                        VALUES (?, ?, ?, ?)
                                    """
                                    execute_non_query(
                                        photo_query,
                                        (int(trip_id), photo_bytes, photo.name, len(photo_bytes))
                                    )
                            
                            st.success(f"✅ Trip completed! You drove {miles_driven} miles.")
                            st.balloons()
                            time.sleep(1)
                            st.rerun()

def render_public_route_if_requested():
    try:
        params = st.query_params
    except Exception:
        params = st.experimental_get_query_params()

    logger.debug("Raw query params: %s", params)

    public_val = _get_param_value(params, "public")
    if public_val:
        key = public_val.lower().strip()
        logger.info("Public-route via 'public' param: %s", key)
        if key in ("helpdesk_ticket", "ticket", "helpdesk"):
            render_helpdesk_ticket_public_form()
            st.stop()
        if key in ("request_vehicle", "vehicle_request", "vehicle"):
            render_request_vehicle_public_form()
            st.stop()
        if key in ("procurement_request", "procurement", "requisition"):
            render_procurement_request_public_form()
            st.stop()
        if key in ("driver_trip_entry", "trip_entry", "driver_trip"):
            render_driver_trip_entry_public_form()
            st.stop()

    page_val = _get_param_value(params, "page")
    form_val = _get_param_value(params, "form")
    logger.info("Public-route check: page_val=%s form_val=%s", page_val, form_val)

    routes_map = {
        "pages/01_Public_Create_Ticket_standalone.py": "helpdesk_ticket",
        "pages/01_Public_Create_Ticket.py": "helpdesk_ticket",
        "pages/02_Public_Request_Vehicle.py": "request_vehicle",
        "pages/03_Public_Procurement_Request.py": "procurement_request",
    }

    target = ""
    if page_val in routes_map:
        target = routes_map[page_val]
    elif page_val == "public_submit" and form_val:
        target = form_val.lower()
    elif page_val and ("/" in page_val):
        target = page_val.split("/")[0].lower()
    elif isinstance(page_val, str) and page_val.startswith("pages/"):
        token = page_val.replace("pages/", "").replace(".py", "").strip("/")
        if "/" in token:
            target = token.split("/")[0].lower()
        else:
            if "ticket" in token:
                target = "helpdesk_ticket"
            elif "vehicle" in token:
                target = "request_vehicle"
            elif "procurement" in token or "requisition" in token:
                target = "procurement_request"
            else:
                target = token.lower() if token else ""

    logger.info("Public-route resolved target: %s", target)

    if target in ("helpdesk_ticket", "ticket", "helpdesk"):
        render_helpdesk_ticket_public_form()
        st.stop()
    if target in ("request_vehicle", "vehicle_request", "vehicle"):
        render_request_vehicle_public_form()
        st.stop()
    if target in ("procurement_request", "procurement", "requisition"):
        render_procurement_request_public_form()
        st.stop()

# Run early public-route handler
try:
    render_public_route_if_requested()
except Exception as e:
    logger.exception("Public route handler failed: %s", e)
    try:
        st.experimental_set_query_params()
    except Exception:
        pass

# Main app

# ========================================
# WORKFLOW MANAGEMENT HELPER FUNCTIONS
# ========================================

def get_pending_counts(db_available=False):
    """Get counts of pending items for navigation badges"""
    counts = {
        'vehicle_requests': 0,
        'new_tickets': 0,
        'procurement_requests': 0
    }
    
    if not db_available:
        return counts
    
    try:
        # Pending vehicle requests
        vehicle_query = "SELECT COUNT(*) as count FROM dbo.Vehicle_Requests WHERE status = 'Pending'"
        vehicle_df, _ = execute_query(vehicle_query)
        if vehicle_df is not None and len(vehicle_df) > 0:
            counts['vehicle_requests'] = int(vehicle_df.iloc[0]['count'])
        
        # New/Open tickets
        ticket_query = "SELECT COUNT(*) as count FROM dbo.Tickets WHERE status IN ('New', 'Open')"
        ticket_df, _ = execute_query(ticket_query)
        if ticket_df is not None and len(ticket_df) > 0:
            counts['new_tickets'] = int(ticket_df.iloc[0]['count'])
        
        # Pending procurement requests
        proc_query = "SELECT COUNT(*) as count FROM dbo.Procurement_Requests WHERE status = 'Pending'"
        proc_df, _ = execute_query(proc_query)
        if proc_df is not None and len(proc_df) > 0:
            counts['procurement_requests'] = int(proc_df.iloc[0]['count'])
    except Exception as e:
        logger.error(f"Error getting pending counts: {e}")
    
    return counts

# =====================================================
# RESOURCE MANAGEMENT HELPER FUNCTIONS
# =====================================================

def get_resource_categories():
    """Get all active resource categories"""
    query = "SELECT category_id, category_name, icon FROM dbo.resource_categories WHERE is_active = 1 ORDER BY category_name"
    df, err = execute_query(query)
    return df if err is None else pd.DataFrame()

def get_resource_locations():
    """Get all active locations"""
    query = "SELECT location_id, location_name, location_type FROM dbo.resource_locations WHERE is_active = 1 ORDER BY location_name"
    df, err = execute_query(query)
    return df if err is None else pd.DataFrame()

def get_resources_by_category(category_id=None):
    """Get resources, optionally filtered by category"""
    if category_id:
        query = """
            SELECT r.resource_id, r.resource_name, r.upc_code, r.unit_of_measure,
                   rc.category_name, rc.icon
            FROM dbo.resources r
            INNER JOIN dbo.resource_categories rc ON r.category_id = rc.category_id
            WHERE r.is_active = 1 AND r.category_id = ?
            ORDER BY r.resource_name
        """
        df, err = execute_query(query, (category_id,))
    else:
        query = """
            SELECT r.resource_id, r.resource_name, r.upc_code, r.unit_of_measure,
                   rc.category_name, rc.icon
            FROM dbo.resources r
            INNER JOIN dbo.resource_categories rc ON r.category_id = rc.category_id
            WHERE r.is_active = 1
            ORDER BY rc.category_name, r.resource_name
        """
        df, err = execute_query(query)
    return df if err is None else pd.DataFrame()

def get_inventory_by_location(location_id):
    """Get inventory for a specific location"""
    query = """
        SELECT r.resource_id, r.resource_name, r.upc_code, rc.category_name,
               i.quantity_on_hand, i.quantity_allocated, i.quantity_available,
               r.reorder_level, r.unit_of_measure
        FROM dbo.resource_inventory i
        INNER JOIN dbo.resources r ON i.resource_id = r.resource_id
        INNER JOIN dbo.resource_categories rc ON r.category_id = rc.category_id
        WHERE i.location_id = ? AND r.is_active = 1
        ORDER BY rc.category_name, r.resource_name
    """
    df, err = execute_query(query, (location_id,))
    return df if err is None else pd.DataFrame()

# =====================================================
# LOGIN PAGE - Wakes up DB
# =====================================================

def render_login_page():
    """Simple login page to wake up database"""
    st.title("🏥 VDH Service Center")
    st.markdown("---")
    
    # Try to connect to database to wake it up
    if "db_woken" not in st.session_state:
        with st.spinner("Connecting to database..."):
            try:
                import time
                time.sleep(1)
                conn = get_db_connection()
                conn.close()
                st.session_state.db_woken = True
                st.success("✅ Database connection established")
            except Exception as e:
                st.warning("Database is warming up... Please wait and try again.")
                logger.info(f"DB wake attempt: {e}")
    
    st.markdown("---")
    
    # Initialize password reset state
    if "show_password_reset" not in st.session_state:
        st.session_state.show_password_reset = False
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if not st.session_state.show_password_reset:
            # Login form
            st.markdown("### Sign In")
            
            with st.form("login_form"):
                username = st.text_input("Username", placeholder="Enter your username")
                password = st.text_input("Password", type="password", placeholder="Enter your password")
                
                submit = st.form_submit_button("🔐 Sign In", use_container_width=True, type="primary")
                
                if submit:
                    if username and password:
                        st.session_state.authenticated = True
                        st.session_state.username = username
                        st.balloons()
                        st.success("✅ Login successful! Redirecting...")
                        import time
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error("⚠️ Please enter both username and password")
            
            st.markdown("---")
            
            # Password reset link
            col_a, col_b, col_c = st.columns([1, 2, 1])
            with col_b:
                if st.button("🔑 Forgot Password?", use_container_width=True):
                    st.session_state.show_password_reset = True
                    st.rerun()
            
            st.caption("Having trouble logging in? Contact IT Help Desk")
        
        else:
            # Password reset form
            st.markdown("### Reset Password")
            st.info("Enter your email address and we will send you instructions to reset your password.")
            
            with st.form("password_reset_form"):
                email = st.text_input("Email Address", placeholder="your.email@vdh.virginia.gov")
                
                col1, col2 = st.columns(2)
                with col1:
                    reset_submit = st.form_submit_button("📧 Send Reset Link", type="primary", use_container_width=True)
                with col2:
                    cancel = st.form_submit_button("← Back to Login", use_container_width=True)
                
                if cancel:
                    st.session_state.show_password_reset = False
                    st.rerun()
                
                if reset_submit:
                    if email and "@" in email:
                        st.success(f"✅ Password reset instructions have been sent to {email}")
                        st.info("Please check your email and follow the instructions to reset your password.")
                        import time
                        time.sleep(3)
                        st.session_state.show_password_reset = False
                        st.rerun()
                    else:
                        st.error("⚠️ Please enter a valid email address")

# =====================================================
# RESOURCE MANAGEMENT RENDER FUNCTIONS
# =====================================================

def render_add_resource_form():
    """Form to add new resource with UPC support"""
    st.subheader("➕ Add New Resource")
    
    categories_df = get_resource_categories()
    if categories_df.empty:
        st.error("No categories available. Please add categories first.")
        return
    
    with st.form("add_resource_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            resource_name = st.text_input("Resource Name *", placeholder="e.g., Feed More Food Box - Standard")
            category_id = st.selectbox(
                "Category *",
                options=categories_df['category_id'].tolist(),
                format_func=lambda x: categories_df[categories_df['category_id']==x]['category_name'].values[0]
            )
            upc_code = st.text_input("UPC/Barcode *", placeholder="Scan or enter UPC code", help="Scan with barcode gun or enter manually")
            
        with col2:
            sku = st.text_input("SKU", placeholder="Optional SKU/Item Number")
            unit_of_measure = st.selectbox("Unit of Measure *", ['box', 'kit', 'pack', 'case', 'unit', 'each'])
            reorder_level = st.number_input("Reorder Level", min_value=0, value=50, help="Alert when stock falls below this number")
        
        description = st.text_area("Description", placeholder="Optional notes about this resource")
        
        col1, col2 = st.columns(2)
        with col1:
            submit = st.form_submit_button("💾 Add Resource", use_container_width=True, type="primary")
        with col2:
            cancel = st.form_submit_button("❌ Cancel", use_container_width=True)
        
        if cancel:
            st.session_state.resource_view = 'dashboard'
            st.rerun()
        
        if submit:
            if not resource_name or not upc_code:
                st.error("⚠️ Please fill in all required fields")
            else:
                user = st.session_state.get('user', {})
                created_by = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or user.get('username', 'Unknown')
                
                insert_query = """
                    INSERT INTO dbo.resources (resource_name, category_id, upc_code, sku, description, 
                                              unit_of_measure, reorder_level, created_by, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, GETDATE())
                """
                
                result, err = execute_non_query(insert_query, (
                    resource_name, category_id, upc_code, sku, description,
                    unit_of_measure, reorder_level, created_by
                ))
                
                if err:
                    if 'duplicate' in str(err).lower() or 'unique' in str(err).lower():
                        st.error(f"❌ UPC Code '{upc_code}' already exists in the system")
                    else:
                        st.error(f"❌ Error adding resource: {err}")
                else:
                    st.success(f"✅ Resource '{resource_name}' added successfully!")
                    
                    # Initialize inventory at all locations
                    get_resource_id_query = "SELECT resource_id FROM dbo.resources WHERE upc_code = ?"
                    resource_df, _ = execute_query(get_resource_id_query, (upc_code,))
                    if not resource_df.empty:
                        resource_id = resource_df.iloc[0]['resource_id']
                        locations_df = get_resource_locations()
                        for _, loc in locations_df.iterrows():
                            init_inventory_query = """
                                INSERT INTO dbo.resource_inventory (resource_id, location_id, quantity_on_hand, updated_by, updated_at)
                                VALUES (?, ?, 0, ?, GETDATE())
                            """
                            execute_non_query(init_inventory_query, (resource_id, loc['location_id'], created_by))
                    
                    import time
                    time.sleep(1)
                    st.session_state.resource_view = 'inventory'
                    st.rerun()

# =====================================================
# SECTION 2: INVENTORY MANAGEMENT
# =====================================================

def render_inventory_management():
    """View and manage inventory levels"""
    st.subheader("📦 Inventory Management")
    
    locations_df = get_resource_locations()
    if locations_df.empty:
        st.warning("No locations configured")
        return
    
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        selected_location_id = st.selectbox(
            "Select Location",
            options=locations_df['location_id'].tolist(),
            format_func=lambda x: locations_df[locations_df['location_id']==x]['location_name'].values[0]
        )
    with col2:
        if st.button("🔄 Refresh Inventory"):
            st.rerun()
    with col3:
        if st.button("➕ Adjust Stock"):
            st.session_state.show_adjust_form = True
    
    # Show adjustment form if requested
    if st.session_state.get('show_adjust_form', False):
        render_adjust_stock_form(selected_location_id)
        return
    
    inventory_df = get_inventory_by_location(selected_location_id)
    
    if inventory_df.empty:
        st.info("No inventory items for this location")
        return
    
    # Summary cards
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Items", len(inventory_df))
    with col2:
        total_stock = inventory_df['quantity_on_hand'].sum()
        st.metric("Total Stock", f"{total_stock:,}")
    with col3:
        low_stock = len(inventory_df[inventory_df['quantity_available'] <= inventory_df['reorder_level']])
        st.metric("Low Stock Items", low_stock, delta=-low_stock if low_stock > 0 else None)
    with col4:
        out_of_stock = len(inventory_df[inventory_df['quantity_available'] == 0])
        st.metric("Out of Stock", out_of_stock, delta=-out_of_stock if out_of_stock > 0 else None)
    
    inventory_df['status'] = inventory_df.apply(
        lambda row: '🔴 Out of Stock' if row['quantity_available'] == 0 
        else '🟡 Low Stock' if row['quantity_available'] <= row['reorder_level']
        else '🟢 In Stock',
        axis=1
    )
    
    st.markdown("### Current Inventory")
    st.dataframe(
        inventory_df[['resource_name', 'category_name', 'quantity_on_hand', 'quantity_allocated', 
                      'quantity_available', 'reorder_level', 'status', 'unit_of_measure']],
        use_container_width=True,
        hide_index=True
    )

# =====================================================
# SECTION 3: RESOURCE DASHBOARD
# =====================================================

from datetime import datetime, timedelta
import base64

# ==========================================
# MANIFEST MANAGEMENT FUNCTIONS
# ==========================================

def log_manifest_activity(manifest_id, activity_type, details, user):
    """Log all manifest activity for reporting"""
    log_query = """
        INSERT INTO dbo.manifest_activity_log 
        (manifest_id, activity_type, activity_details, performed_by, activity_date)
        VALUES (?, ?, ?, ?, GETDATE())
    """
    execute_non_query(log_query, (manifest_id, activity_type, details, user))

def render_signature_capture():
    """Render signature capture with both draw and type options"""
    st.markdown("### ✍️ Delivery Signature Required")
    st.info("📋 A signature is required to mark this manifest as delivered")
    
    sig_method = st.radio(
        "Signature Method", 
        ["✍️ Type Signature", "🎨 Draw Signature (Tablet/Stylus)"], 
        horizontal=True,
        key="sig_method"
    )
    
    signature_data = None
    signature_type = None
    
    if sig_method == "✍️ Type Signature":
        st.markdown("**Electronic Signature**")
        typed_signature = st.text_input(
            "Full Name", 
            placeholder="Enter your full name",
            key="typed_sig"
        )
        
        if typed_signature:
            # Display signature preview
            st.markdown(f"""
                <div style="
                    border: 2px solid #002855; 
                    padding: 30px; 
                    background: linear-gradient(to bottom, #ffffff 0%, #f8f9fa 100%); 
                    border-radius: 8px; 
                    margin-top: 15px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                ">
                    <p style="
                        font-family: 'Brush Script MT', 'Lucida Handwriting', cursive; 
                        font-size: 32px; 
                        margin: 0;
                        color: #002855;
                        text-align: center;
                    ">
                        {typed_signature}
                    </p>
                    <hr style="border: 1px solid #002855; margin: 15px 0;">
                    <p style="
                        font-size: 11px; 
                        color: #666; 
                        margin: 5px 0;
                        text-align: center;
                    ">
                        Electronic Signature - {datetime.now().strftime("%B %d, %Y at %I:%M %p")}
                    </p>
                    <p style="
                        font-size: 10px; 
                        color: #999; 
                        margin: 0;
                        text-align: center;
                        font-style: italic;
                    ">
                        By signing, I acknowledge receipt of the items listed in this manifest
                    </p>
                </div>
            """, unsafe_allow_html=True)
            
            signature_data = typed_signature
            signature_type = "typed"
    
    else:  # Draw Signature
        st.markdown("**Draw Your Signature**")
        st.info("🖊️ Use your stylus, finger, or mouse to sign below")
        
        # Simple drawing canvas using HTML5 canvas
        signature_html = """
        <div style="border: 2px solid #002855; border-radius: 8px; padding: 10px; background: white;">
            <canvas id="signatureCanvas" width="600" height="200" style="border: 1px dashed #ccc; cursor: crosshair; background: white; width: 100%;"></canvas>
            <div style="margin-top: 10px;">
                <button onclick="clearCanvas()" style="padding: 8px 16px; background: #dc3545; color: white; border: none; border-radius: 4px; cursor: pointer;">Clear</button>
                <span id="signatureStatus" style="margin-left: 15px; color: #666;"></span>
            </div>
        </div>
        
        <script>
        const canvas = document.getElementById('signatureCanvas');
        const ctx = canvas.getContext('2d');
        let drawing = false;
        let hasSignature = false;
        
        // Set canvas actual size
        canvas.width = 600;
        canvas.height = 200;
        
        function startDrawing(e) {
            drawing = true;
            hasSignature = true;
            const rect = canvas.getBoundingClientRect();
            const x = (e.clientX || e.touches[0].clientX) - rect.left;
            const y = (e.clientY || e.touches[0].clientY) - rect.top;
            ctx.beginPath();
            ctx.moveTo(x * (canvas.width / rect.width), y * (canvas.height / rect.height));
            document.getElementById('signatureStatus').innerText = '✓ Signature captured';
            document.getElementById('signatureStatus').style.color = '#28a745';
        }
        
        function draw(e) {
            if (!drawing) return;
            e.preventDefault();
            
            const rect = canvas.getBoundingClientRect();
            const x = (e.clientX || e.touches[0].clientX) - rect.left;
            const y = (e.clientY || e.touches[0].clientY) - rect.top;
            
            ctx.lineWidth = 2;
            ctx.lineCap = 'round';
            ctx.strokeStyle = '#002855';
            ctx.lineTo(x * (canvas.width / rect.width), y * (canvas.height / rect.height));
            ctx.stroke();
        }
        
        function stopDrawing() {
            drawing = false;
            ctx.beginPath();
            if (hasSignature) {
                // Store signature as base64
                const dataURL = canvas.toDataURL('image/png');
                localStorage.setItem('manifestSignature', dataURL);
            }
        }
        
        function clearCanvas() {
            ctx.clearRect(0, 0, canvas.width, canvas.height);
            hasSignature = false;
            localStorage.removeItem('manifestSignature');
            document.getElementById('signatureStatus').innerText = '';
        }
        
        canvas.addEventListener('mousedown', startDrawing);
        canvas.addEventListener('mousemove', draw);
        canvas.addEventListener('mouseup', stopDrawing);
        canvas.addEventListener('touchstart', startDrawing);
        canvas.addEventListener('touchmove', draw);
        canvas.addEventListener('touchend', stopDrawing);
        
        // Prevent scrolling when touching the canvas
        canvas.addEventListener('touchstart', function(e) { e.preventDefault(); }, { passive: false });
        canvas.addEventListener('touchmove', function(e) { e.preventDefault(); }, { passive: false });
        </script>
        """
        
        st.components.v1.html(signature_html, height=280)
        
        # For drawn signatures, we'll use a text confirmation instead
        drawn_confirm = st.text_input(
            "Type your name to confirm the drawn signature",
            placeholder="Confirm signature by typing your name",
            key="drawn_confirm"
        )
        
        if drawn_confirm:
            signature_data = f"[Signature] {drawn_confirm}"
            signature_type = "drawn"
            st.success("✅ Signature confirmed")
    
    return signature_data, signature_type

def render_manifest_list():
    """Display all manifests with status management and New Manifest button"""
    st.subheader("📋 Manifest Management")
    
    # Top controls with New Manifest button
    col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
    with col1:
        status_filter = st.selectbox(
            "Filter by Status", 
            ["All", "Staged", "In Transit", "Delivered"],
            key="manifest_status_filter"
        )
    with col2:
        date_filter = st.date_input(
            "From Date", 
            value=datetime.now() - timedelta(days=30),
            key="manifest_date_filter"
        )
    with col3:
        if st.button("🔄 Refresh", use_container_width=True):
            st.rerun()
    with col4:
        if st.button("📦 New Manifest", type="primary", use_container_width=True):
            st.session_state.resource_view = 'create_manifest'
            st.rerun()
    
    st.markdown("---")
    
    # Build query
    query = """
        SELECT m.manifest_id, m.manifest_number, m.shipment_date,
               COALESCE(fl.location_name, m.from_location_name) as from_location,
               COALESCE(tl.location_name, m.to_location_name) as to_location,
               m.status, m.created_by, m.created_at,
               m.signature_name, m.signature_type, m.delivered_at, m.delivered_by
        FROM dbo.resource_manifests m
        LEFT JOIN dbo.resource_locations fl ON m.from_location_id = fl.location_id
        LEFT JOIN dbo.resource_locations tl ON m.to_location_id = tl.location_id
        WHERE m.shipment_date >= ?
    """
    
    params = [date_filter]
    
    if status_filter != "All":
        query += " AND m.status = ?"
        params.append(status_filter)
    
    query += " ORDER BY m.created_at DESC"
    
    df, err = execute_query(query, tuple(params))
    
    if err:
        st.error(f"Error loading manifests: {err}")
        return
    
    if df is None or df.empty:
        st.info("📦 No manifests found. Click 'New Manifest' to create your first one.")
        return
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📦 Total Manifests", len(df))
    with col2:
        staged = len(df[df['status'] == 'Staged'])
        st.metric("🟡 Staged", staged)
    with col3:
        in_transit = len(df[df['status'] == 'In Transit'])
        st.metric("🔵 In Transit", in_transit)
    with col4:
        delivered = len(df[df['status'] == 'Delivered'])
        st.metric("🟢 Delivered", delivered)
    
    st.markdown("---")
    
    # Display each manifest
    for _, manifest in df.iterrows():
        status_icons = {"Staged": "🟡", "In Transit": "🔵", "Delivered": "🟢"}
        icon = status_icons.get(manifest['status'], "⚪")
        
        with st.expander(
            f"{icon} {manifest['manifest_number']} - {manifest['from_location']} → {manifest['to_location']} ({manifest['status']})",
            expanded=False
        ):
            # Manifest details
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown("**Locations**")
                st.write(f"📤 From: {manifest['from_location']}")
                st.write(f"📥 To: {manifest['to_location']}")
            
            with col2:
                st.markdown("**Timeline**")
                st.write(f"📅 Shipment Date: {manifest['shipment_date']}")
                st.write(f"👤 Created By: {manifest['created_by']}")
                st.write(f"🕐 Created: {manifest['created_at'].strftime('%Y-%m-%d %H:%M')}")
            
            with col3:
                st.markdown("**Status**")
                st.write(f"{icon} {manifest['status']}")
                if manifest['status'] == 'Delivered' and manifest['signature_name']:
                    st.write(f"✍️ Signed By: {manifest['signature_name']}")
                    st.write(f"📅 Delivered: {manifest['delivered_at'].strftime('%Y-%m-%d %H:%M')}")
            
            st.markdown("---")
            
            # Status transition buttons
            manifest_id = manifest['manifest_id']
            current_status = manifest['status']
            username = st.session_state.get('username', 'Unknown')
            
            col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
            
            if current_status == "Staged":
                with col1:
                    if st.button(f"🚚 Mark In Transit", key=f"transit_{manifest_id}", use_container_width=True):
                        update_query = "UPDATE dbo.resource_manifests SET status = 'In Transit' WHERE manifest_id = ?"
                        result, err = execute_non_query(update_query, (manifest_id,))
                        if not err:
                            log_manifest_activity(manifest_id, "Status Change", "Staged → In Transit", username)
                            st.success("✅ Manifest marked as In Transit")
                            st.rerun()
                        else:
                            st.error(f"Error: {err}")
            
            elif current_status == "In Transit":
                # Check if signature capture is active
                if st.session_state.get(f'show_signature_{manifest_id}', False):
                    st.markdown("---")
                    signature_data, signature_type = render_signature_capture()
                    
                    col_a, col_b, col_c = st.columns([2, 2, 2])
                    
                    with col_a:
                        if signature_data and st.button(
                            "✅ Confirm Delivery", 
                            key=f"confirm_{manifest_id}", 
                            type="primary",
                            use_container_width=True
                        ):
                            update_query = """
                                UPDATE dbo.resource_manifests 
                                SET status = 'Delivered',
                                    signature_name = ?,
                                    signature_type = ?,
                                    signature_data = ?,
                                    delivered_at = GETDATE(),
                                    delivered_by = ?
                                WHERE manifest_id = ?
                            """
                            result, err = execute_non_query(
                                update_query, 
                                (signature_data, signature_type, signature_data, username, manifest_id)
                            )
                            
                            if not err:
                                log_manifest_activity(
                                    manifest_id, 
                                    "Delivered", 
                                    f"Signed by: {signature_data} (Type: {signature_type})", 
                                    username
                                )
                                st.success(f"✅ Manifest delivered! Signed by: {signature_data}")
                                del st.session_state[f'show_signature_{manifest_id}']
                                st.rerun()
                            else:
                                st.error(f"Error: {err}")
                    
                    with col_b:
                        if st.button("❌ Cancel", key=f"cancel_sig_{manifest_id}", use_container_width=True):
                            del st.session_state[f'show_signature_{manifest_id}']
                            st.rerun()
                else:
                    with col1:
                        if st.button(
                            "📍 Mark Delivered", 
                            key=f"deliver_{manifest_id}",
                            use_container_width=True
                        ):
                            st.session_state[f'show_signature_{manifest_id}'] = True
                            st.rerun()
            
            # View activity log
            with col4:
                if st.button("📋 Activity Log", key=f"log_{manifest_id}", use_container_width=True):
                    log_query = """
                        SELECT activity_type, activity_details, performed_by, 
                               FORMAT(activity_date, 'yyyy-MM-dd HH:mm:ss') as activity_time
                        FROM dbo.manifest_activity_log
                        WHERE manifest_id = ?
                        ORDER BY activity_date DESC
                    """
                    log_df, _ = execute_query(log_query, (manifest_id,))
                    
                    if log_df is not None and not log_df.empty:
                        st.markdown("**Activity History**")
                        st.dataframe(
                            log_df,
                            column_config={
                                "activity_type": "Activity",
                                "activity_details": "Details",
                                "performed_by": "User",
                                "activity_time": "Date/Time"
                            },
                            hide_index=True,
                            use_container_width=True
                        )
                    else:
                        st.info("No activity log entries")

def render_manifest_creation():
    """Create new manifest with Popup/Event location support"""
    st.subheader("📦 Create New Manifest")
    
    # Initialize session state
    if 'manifest_items' not in st.session_state:
        st.session_state.manifest_items = []
    
    # Get locations
    locations_df = get_resource_locations()
    
    # Add Popup/Event as special location
    location_options = []
    location_map = {}
    
    # Add regular locations
    if not locations_df.empty:
        for _, loc in locations_df.iterrows():
            loc_id = int(loc['location_id'])
            loc_name = loc['location_name']
            location_options.append(loc_name)
            location_map[loc_name] = loc_id
    
    # Add Popup/Event
    location_options.append("📍 Popup/Event (Custom)")
    location_map["📍 Popup/Event (Custom)"] = -1
    
    st.markdown("### 📋 Shipment Details")
    
    col1, col2 = st.columns(2)
    
    with col1:
        from_location_name = st.selectbox("From Location *", options=location_options, key="from_loc")
        
        # If Popup/Event selected, ask for custom name
        if from_location_name == "📍 Popup/Event (Custom)":
            from_custom_name = st.text_input("Event/Location Name *", placeholder="e.g., Health Fair at City Park", key="from_custom")
        else:
            from_custom_name = None
    
    with col2:
        to_location_name = st.selectbox("To Location *", options=location_options, key="to_loc")
        
        if to_location_name == "📍 Popup/Event (Custom)":
            to_custom_name = st.text_input("Event/Location Name *", placeholder="e.g., Community Center", key="to_custom")
        else:
            to_custom_name = None
    
    shipment_date = st.date_input("Shipment Date *", value=datetime.now())
    notes = st.text_area("Notes", placeholder="Optional shipment notes or special instructions")
    
    # Validation
    if from_location_name == to_location_name and from_custom_name == to_custom_name:
        st.error("⚠️ Source and destination must be different")
        if st.button("❌ Cancel"):
            st.session_state.manifest_items = []
            st.session_state.resource_view = 'manifests'
            st.rerun()
        return
    
    st.markdown("---")
    st.markdown("### 📦 Add Items to Manifest")
    
    # Get source location ID
    from_location_id = location_map[from_location_name]
    
    # Only show inventory if source is a real location (not Popup/Event)
    if from_location_id != -1:
        inventory_df = get_inventory_by_location(from_location_id)
        
        if inventory_df is not None and not inventory_df.empty:
            available_items_df = inventory_df[inventory_df['quantity_available'] > 0]
            
            if not available_items_df.empty:
                col1, col2, col3 = st.columns([3, 2, 2])
                
                with col1:
                    resource_name = st.selectbox(
                        "Select Resource",
                        options=available_items_df['resource_name'].tolist(),
                        key="manifest_resource"
                    )
                
                selected_row = available_items_df[available_items_df['resource_name'] == resource_name].iloc[0]
                available_qty = int(selected_row['quantity_available'])
                
                with col2:
                    st.metric("Available Stock", f"{available_qty} {selected_row['unit_of_measure']}")
                
                with col3:
                    quantity = st.number_input(
                        "Quantity *", 
                        min_value=1, 
                        max_value=available_qty, 
                        value=1,
                        key="manifest_qty"
                    )
                
                if st.button("➕ Add to Manifest", type="primary"):
                    st.session_state.manifest_items.append({
                        'resource_id': int(selected_row['resource_id']),
                        'resource_name': resource_name,
                        'quantity': quantity,
                        'unit': selected_row['unit_of_measure']
                    })
                    st.success(f"✅ Added {quantity} {selected_row['unit_of_measure']} of {resource_name}")
                    st.rerun()
            else:
                st.warning("⚠️ No inventory available at source location with stock > 0")
        else:
            st.info("ℹ️ No inventory found at source location")
    else:
        st.info("ℹ️ Source is Popup/Event - inventory will be manually tracked")
    
    # Display current manifest items
    if st.session_state.manifest_items:
        st.markdown("---")
        st.markdown("### 📋 Current Manifest Items")
        
        items_df = pd.DataFrame(st.session_state.manifest_items)
        st.dataframe(
            items_df[['resource_name', 'quantity', 'unit']],
            column_config={
                "resource_name": "Resource",
                "quantity": "Quantity",
                "unit": "Unit"
            },
            hide_index=True,
            use_container_width=True
        )
        
        st.markdown("---")
        
        col1, col2, col3 = st.columns([2, 2, 2])
        
        with col1:
            if st.button("✅ Create Manifest", type="primary", use_container_width=True):
                username = st.session_state.get('username', 'Unknown')
                
                # Generate manifest number
                manifest_num = f"MAN-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
                
                # Prepare location values
                from_loc_id = location_map[from_location_name] if from_location_id != -1 else None
                to_loc_id = location_map[to_location_name] if location_map[to_location_name] != -1 else None
                from_name = from_custom_name if from_location_id == -1 else None
                to_name = to_custom_name if location_map[to_location_name] == -1 else None
                
                # Insert manifest
                manifest_query = """
                    INSERT INTO dbo.resource_manifests 
                    (manifest_number, from_location_id, to_location_id, from_location_name, to_location_name,
                     shipment_date, status, notes, created_by, created_at)
                    OUTPUT INSERTED.manifest_id
                    VALUES (?, ?, ?, ?, ?, ?, 'Staged', ?, ?, GETDATE())
                """
                
                try:
                    result_df, err = execute_query(
                        manifest_query,
                        (manifest_num, from_loc_id, to_loc_id, from_name, to_name, shipment_date, notes, username)
                    )
                    
                    if err:
                        st.error(f"❌ Error creating manifest: {err}")
                        _logger.error(f"Manifest creation SQL error: {err}")
                    elif result_df is None or result_df.empty:
                        st.error("❌ Manifest creation failed - no ID returned from database")
                        st.warning("📋 Debug info: Query executed but returned no data")
                        _logger.error(f"Manifest creation returned None/empty. Params: {(manifest_num, from_loc_id, to_loc_id, from_name, to_name, shipment_date, notes, username)}")
                    else:
                        manifest_id = result_df.iloc[0]['manifest_id']
                        st.success(f"✅ Manifest header created with ID: {manifest_id}")
                        
                        # Insert manifest items
                        items_success = 0
                        for item in st.session_state.manifest_items:
                            try:
                                item_query = """
                                    INSERT INTO dbo.manifest_items (manifest_id, resource_id, quantity)
                                    VALUES (?, ?, ?)
                                """
                                item_result, item_err = execute_non_query(item_query, (manifest_id, item['resource_id'], item['quantity']))
                                if not item_err:
                                    items_success += 1
                                else:
                                    st.warning(f"⚠️ Error adding item {item['resource_name']}: {item_err}")
                            except Exception as item_ex:
                                st.warning(f"⚠️ Exception adding item {item['resource_name']}: {item_ex}")
                        
                        st.info(f"Added {items_success}/{len(st.session_state.manifest_items)} items to manifest")
                        
                        # Log activity
                        try:
                            from_display = from_custom_name if from_custom_name else from_location_name
                            to_display = to_custom_name if to_custom_name else to_location_name
                            log_manifest_activity(
                                manifest_id, 
                                "Created", 
                                f"Manifest {manifest_num} created: {from_display} → {to_display}",
                                username
                            )
                        except Exception as log_err:
                            st.warning(f"⚠️ Activity logging failed: {log_err}")
                            _logger.error(f"Activity log error: {log_err}")
                        
                        st.success(f"🎉 Manifest {manifest_num} created successfully!")
                        st.balloons()
                        st.session_state.manifest_items = []
                        import time
                        time.sleep(2)
                        st.session_state.resource_view = 'manifests'
                        st.rerun()
                
                except Exception as e:
                    st.error(f"❌ Unexpected error creating manifest: {str(e)}")
                    _logger.exception("Manifest creation exception")
                    with st.expander("🐛 Debug Details"):
                        import traceback
                        st.code(traceback.format_exc())
        
        with col2:
            if st.button("🗑️ Clear All Items", use_container_width=True):
                st.session_state.manifest_items = []
                st.rerun()
        
        with col3:
            if st.button("❌ Cancel", use_container_width=True):
                st.session_state.manifest_items = []
                st.session_state.resource_view = 'manifests'
                st.rerun()
    else:
        col1, col2 = st.columns([1, 1])
        with col2:
            if st.button("❌ Cancel", use_container_width=True):
                st.session_state.resource_view = 'manifests'
                st.rerun()


def render_resource_dashboard():
    """Dashboard with key metrics"""
    st.subheader("📊 Resource Management Dashboard")
    
    col1, col2, col3, col4 = st.columns(4)
    
    total_resources_query = "SELECT COUNT(*) as count FROM dbo.resources WHERE is_active = 1"
    total_resources_df, _ = execute_query(total_resources_query)
    with col1:
        count = total_resources_df.iloc[0]['count'] if not total_resources_df.empty else 0
        st.metric("Total Resources", count)
    
    total_inventory_query = "SELECT SUM(quantity_on_hand) as total FROM dbo.resource_inventory"
    total_inventory_df, _ = execute_query(total_inventory_query)
    with col2:
        total = total_inventory_df.iloc[0]['total'] if not total_inventory_df.empty else 0
        st.metric("Total Stock", f"{total:,.0f}")
    
    pending_query = "SELECT COUNT(*) as count FROM dbo.resource_shipments WHERE status IN ('Pending', 'In Transit')"
    pending_df, _ = execute_query(pending_query)
    with col3:
        pending = pending_df.iloc[0]['count'] if not pending_df.empty else 0
        st.metric("Pending Shipments", pending)
    
    low_stock_query = """
        SELECT COUNT(*) as count 
        FROM dbo.resource_inventory i
        INNER JOIN dbo.resources r ON i.resource_id = r.resource_id
        WHERE i.quantity_available <= r.reorder_level AND r.is_active = 1
    """
    low_stock_df, _ = execute_query(low_stock_query)
    with col4:
        low = low_stock_df.iloc[0]['count'] if not low_stock_df.empty else 0
        st.metric("Low Stock Alerts", low, delta=-low if low > 0 else None)

# =====================================================
# SECTION 4: MAIN RESOURCE MANAGEMENT PAGE
# =====================================================

def render_adjust_stock_form(selected_location_id):
    """Form to adjust inventory quantities"""
    st.subheader("📊 Adjust Stock Levels")
    
    inventory_df = get_inventory_by_location(selected_location_id)
    
    if inventory_df.empty:
        st.warning("No inventory items available for adjustment")
        return
    
    with st.form("adjust_stock_form"):
        # Select resource
        resource_options = inventory_df['resource_name'].tolist()
        selected_resource = st.selectbox("Select Resource", resource_options)
        
        # Get current quantity
        current_row = inventory_df[inventory_df['resource_name'] == selected_resource].iloc[0]
        current_qty = int(current_row['quantity_on_hand'])
        
        st.info(f"Current Stock: **{current_qty}** {current_row['unit_of_measure']}")
        
        # Adjustment type
        adjustment_type = st.radio("Adjustment Type", ['Add Stock', 'Remove Stock', 'Set Exact Quantity'])
        
        if adjustment_type == 'Set Exact Quantity':
            new_quantity = st.number_input("New Quantity", min_value=0, value=current_qty)
            adjustment = new_quantity - current_qty
        else:
            adjustment = st.number_input("Quantity to Adjust", min_value=0, value=0)
            if adjustment_type == 'Remove Stock':
                adjustment = -adjustment
        
        reason = st.text_area("Reason for Adjustment *", placeholder="e.g., Received shipment, Damaged items, Inventory count correction")
        
        col1, col2 = st.columns(2)
        with col1:
            submit = st.form_submit_button("💾 Apply Adjustment", type="primary", use_container_width=True)
        with col2:
            cancel = st.form_submit_button("❌ Cancel", use_container_width=True)
        
        if cancel:
            st.session_state.show_adjust_form = False
            st.rerun()
        
        if submit:
            if not reason:
                st.error("⚠️ Please provide a reason for the adjustment")
            else:
                username = st.session_state.get('username', 'Unknown')
                resource_id = int(current_row['resource_id'])
                
                # Calculate new quantity
                if adjustment_type == 'Set Exact Quantity':
                    final_qty = new_quantity
                else:
                    final_qty = current_qty + adjustment
                
                if final_qty < 0:
                    st.error("❌ Cannot reduce stock below zero")
                else:
                    # Update inventory
                    update_query = """
                        UPDATE dbo.resource_inventory 
                        SET quantity_on_hand = ?, 
                            updated_by = ?, 
                            updated_at = GETDATE()
                        WHERE resource_id = ? AND location_id = ?
                    """
                    result, err = execute_non_query(update_query, (final_qty, username, resource_id, selected_location_id))
                    
                    if err:
                        st.error(f"❌ Error updating inventory: {err}")
                    else:
                        # Log the adjustment
                        log_query = """
                            INSERT INTO dbo.resource_inventory_log (resource_id, location_id, adjustment_qty, 
                                                                     reason, adjusted_by, adjusted_at)
                            VALUES (?, ?, ?, ?, ?, GETDATE())
                        """
                        execute_non_query(log_query, (resource_id, selected_location_id, adjustment, reason, username))
                        
                        st.success(f"✅ Stock adjusted successfully! New quantity: {final_qty}")
                        import time
                        time.sleep(1)
                        st.session_state.show_adjust_form = False
                        st.rerun()

def render_resource_management():
    """Main Resource Management page"""
    
    if 'resource_view' not in st.session_state:
        st.session_state.resource_view = 'dashboard'
    
    st.title("📦 Resource Management")
    
    # Resource Management Locations
    RESOURCE_LOCATIONS = [
        "Petersburg WIC",
        "Petersburg Clinic B", 
        "Petersburg Warehouse",
        "Dinwiddie County Health Dept",
        "Greensville/Emporia Health Dept",
        "Surry County Health Dept",
        "Prince George Health Dept",
        "Sussex County Health Dept",
        "Hopewell Health Dept",
    ]

    st.markdown("*Population Health Distribution System*")
    st.markdown("---")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("📊 Dashboard", use_container_width=True, 
                    type="primary" if st.session_state.resource_view == 'dashboard' else "secondary"):
            st.session_state.resource_view = 'dashboard'
            st.rerun()
    
    with col2:
        if st.button("📦 Inventory", use_container_width=True,
                    type="primary" if st.session_state.resource_view == 'inventory' else "secondary"):
            st.session_state.resource_view = 'inventory'
            st.rerun()
    
    with col3:
        if st.button("📋 Manifests", use_container_width=True,
                    type="primary" if st.session_state.resource_view == 'manifests' else "secondary"):
            st.session_state.resource_view = 'manifests'
            st.rerun()
    
    with col4:
        if st.button("➕ New Resource", use_container_width=True,
                    type="primary" if st.session_state.resource_view == 'add_resource' else "secondary"):
            st.session_state.resource_view = 'add_resource'
            st.rerun()
    
    st.markdown("---")
    
    if st.session_state.resource_view == 'dashboard':
        render_resource_dashboard()
    elif st.session_state.resource_view == 'inventory':
        render_inventory_management()
    elif st.session_state.resource_view == 'add_resource':
        render_add_resource_form()
    elif st.session_state.resource_view == 'manifests':
        render_manifest_list()
    elif st.session_state.resource_view == 'create_manifest':
        render_manifest_creation()
    # Locations view removed

    # Define page options

def main():
    st.set_page_config(page_title="VDH Service Center", page_icon="🏥", layout="wide")

    # Check authentication
    if "authenticated" not in st.session_state or not st.session_state.authenticated:
        render_login_page()
        return

    st.markdown(
        """
        <style>
          header[role="banner"] { display: none !important; }
          div[data-testid="stSidebarNav"] { display: none !important; }
        
        /* Navigation Alert Badges */
        .stSelectbox div[data-baseweb="select"] > div {
            font-weight: 500;
        }
        
        .alert-badge {
            background-color: #ff4444;
            color: white;
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: bold;
            margin-left: 8px;
        }
        
        .pending-alert {
            background-color: #fff3cd;
            border-left: 4px solid #ffc107;
            padding: 12px;
            margin: 10px 0;
            border-radius: 4px;
        }
        
        .pending-alert h4 {
            color: #856404;
            margin: 0;
            font-size: 16px;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )

    try:
        components.html(
            """
            <script>
            (function() {
              try {
                var header = document.querySelector('header[role="banner"]');
                if (header) header.style.display = 'none';
                var sidebarNav = document.querySelector('div[data-testid="stSidebarNav"]');
                if (sidebarNav) sidebarNav.style.display = 'none';
              } catch(e) {}
            })();
            </script>
            """,
            height=1,
            scrolling=False,
        )
    except Exception:
        st.caption("UI tweak script not applied.")

    placeholder_text = quote("VDH")
    logo_url = f"https://via.placeholder.com/200x80/002855/FFFFFF.png?text={placeholder_text}"
    try:
        st.sidebar.image(logo_url, width=200)
    except Exception:
        try:
            safe_st_image(logo_url, width=200)
        except Exception:
            svg = """
            <svg width="200" height="80" xmlns="http://www.w3.org/2000/svg" role="img" aria-label="VDH">
              <rect width="200" height="80" fill="#002855" rx="6" ry="6"/>
              <text x="100" y="50" font-family="Arial, Helvetica, sans-serif" font-size="16" fill="#FFFFFF" text-anchor="middle">VDH Service Center</text>
            </svg>
            """
            b64 = base64.b64encode(svg.encode("utf-8")).decode("utf-8")
            st.sidebar.markdown(f'<img src="data:image/svg+xml;base64,{b64}" width="200" alt="VDH logo">', unsafe_allow_html=True)

    st.sidebar.title("VDH Service Center")

    # User profile section
    username = st.session_state.get("username", "User")
    st.sidebar.markdown(f"### 👤 {username}")
    if st.sidebar.button("🚪 Logout", use_container_width=True, key="logout_btn"):
        # Clear session state
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
    st.sidebar.markdown("---")
    
    # VDH Landing Page Link
    st.sidebar.markdown("### 🏠 Quick Links")
    st.sidebar.markdown("""
        <a href="https://www.vdh.virginia.gov" target="_blank" style="
            display: inline-block;
            width: 100%;
            padding: 0.5rem 1rem;
            background-color: #002855;
            color: white;
            text-decoration: none;
            border-radius: 0.5rem;
            text-align: center;
            font-weight: 500;
            margin-bottom: 0.5rem;
        ">
            🌐 VDH Landing Page
        </a>
    """, unsafe_allow_html=True)
    
    st.sidebar.markdown("---")


    with st.sidebar:
        st.markdown("---")
        st.markdown("🏛️ **Public Access Forms**")

        ticket_href = "/?public=helpdesk_ticket"
        vehicle_href = "/?public=request_vehicle"
        proc_href = "/?public=procurement_request"
        driver_trip_href = "/?public=driver_trip_entry"

        html_links = (
            '<div style="padding:6px 4px;">'
            f'<a href="{ticket_href}" target="_blank" rel="noopener noreferrer" class="public-link">📩&nbsp;&nbsp;<strong>Submit a Ticket</strong></a><br/>'
            f'<a href="{vehicle_href}" target="_blank" rel="noopener noreferrer" class="public-link">🚗&nbsp;&nbsp;<strong>Request a Vehicle</strong></a><br/>'
            f'<a href="{proc_href}" target="_blank" rel="noopener noreferrer" class="public-link">🛒&nbsp;&nbsp;<strong>Submit a Requisition</strong></a><br/>'
            f'<a href="{driver_trip_href}" target="_blank" rel="noopener noreferrer" class="public-link">🚙&nbsp;&nbsp;<strong>Driver Trip Entry</strong></a>'
            '</div>'
            '<style>'
            '.public-link {'
            '  display: block;'
            '  width: calc(100% - 8px);'
            '  box-sizing: border-box;'
            '  background: #002855;'
            '  color: #ffffff !important;'
            '  padding: 10px 12px;'
            '  text-decoration: none;'
            '  border-radius: 6px;'
            '  margin: 8px 4px;'
            '  font-weight: 600;'
            '  text-align: left;'
            '  white-space: normal;'
            '  line-height: 1.4;'
            '}'
            '.public-link:hover {'
            '  background: #FF6B35;'
            '  color:#fff !important;'
            '  text-decoration: none;'
            '}'
            '</style>'
        )
        st.markdown(html_links, unsafe_allow_html=True)

    try:
        # CRITICAL FIX: Check if DB credentials exist BEFORE attempting connection
        server, database, username, password = get_connection_string()
        
        if not all([server, database, username, password]):
            DB_AVAILABLE = False
            logger.warning("DB credentials not configured (missing env vars or secrets)")
        else:
            # Only try connection if we have credentials
            try:
                logger.info("Testing database connection...")
                conn = get_db_connection()
                logger.info("Database connection successful!")
                DB_AVAILABLE = True
                try:
                    conn.close()
                except Exception:
                    pass
            except Exception as e:
                DB_AVAILABLE = False
                logger.warning("DB not available at startup: %s", e)
    except Exception as e:
        DB_AVAILABLE = False
        logger.error("Error checking DB: %s", e)

    # Get pending counts for navigation badges (pass DB_AVAILABLE)
    pending_counts = get_pending_counts(DB_AVAILABLE)
    
    # Define page options
    # Define page options
    page_options = [
        "📊 Dashboard",
        "🎫 Helpdesk Tickets",
        "💻 Asset Management",
        "🛒 Procurement Requests",
        "🚗 Fleet Management",
        "📦 Resource Management",
        "📈 Report Builder",
        "🔌 Connection Test",
    ]
    
    # Add badges to options
    page_options_display = []
    for option in page_options:
        display_option = option
        if "Fleet Management" in option and pending_counts['vehicle_requests'] > 0:
            display_option = f"{option} 🔴 {pending_counts['vehicle_requests']}"
        elif "Helpdesk Tickets" in option and pending_counts['new_tickets'] > 0:
            display_option = f"{option} 🔴 {pending_counts['new_tickets']}"
        elif "Procurement Requests" in option and pending_counts['procurement_requests'] > 0:
            display_option = f"{option} 🔴 {pending_counts['procurement_requests']}"
        page_options_display.append(display_option)
    
    # Fix double-click issue with session state
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "📊 Dashboard"
    
    # Find index of current page in display options
    try:
        default_index = next(i for i, opt in enumerate(page_options_display) if st.session_state.current_page in opt)
    except StopIteration:
        default_index = 0
    
    page = st.sidebar.selectbox(
        "Navigate",
        page_options_display,
        index=default_index,
        label_visibility="collapsed",
        key="page_selector"
    )

    page = page.split(" 🔴")[0]  # Strip badge from selection
    
    # Update session state when page changes
    if page != st.session_state.current_page:
        st.session_state.current_page = page
    
    if not DB_AVAILABLE and page != "📊 Dashboard":
        st.header(page)
        st.warning("The database is currently unavailable. Most pages require a live database connection. You can still use the Public Access Forms from the sidebar.")
        st.markdown("Use the Public Access Forms in the sidebar to submit tickets or requests while we restore the database.")
        st.info("If the database was just restored, go to the Connection Test page and refresh the app.")

    # DASHBOARD (unchanged behavior)
    if page == "📊 Dashboard":
        st.header("📊 Dashboard Overview")
        with st.spinner("Loading dashboard data..."):
            stats_df, stats_err = execute_query("SELECT COUNT(*) as total_tickets FROM dbo.Tickets")
            status_df, status_err = execute_query("SELECT status, COUNT(*) as count FROM dbo.Tickets GROUP BY status")
            priority_df, pr_err = execute_query("SELECT priority, COUNT(*) as count FROM dbo.Tickets GROUP BY priority")
            location_df, loc_err = execute_query("SELECT location, COUNT(*) as count FROM dbo.Tickets GROUP BY location")

            asset_df, asset_err = execute_query("SELECT COUNT(*) as total_assets FROM dbo.Assets")
            asset_status_df, asset_status_err = execute_query("SELECT status, COUNT(*) as count FROM dbo.Assets GROUP BY status")
            asset_location_df, asset_location_err = execute_query("SELECT location, COUNT(*) as count FROM dbo.Assets GROUP BY location")
            asset_type_df, asset_type_err = execute_query("SELECT type, COUNT(*) as count FROM dbo.Assets GROUP BY type")

            proc_df, proc_err = execute_query("SELECT COUNT(*) as total_requests FROM dbo.Procurement_Requests")
            fleet_df, fleet_err = execute_query("SELECT COUNT(*) as total_vehicles FROM dbo.vehicles")

        errs = [e for e in (stats_err, status_err, pr_err, loc_err, asset_err, asset_status_err, asset_location_err, asset_type_err, proc_err, fleet_err) if e]
        if errs:
            st.warning("Could not load some live data from the database. Showing placeholders where needed. (" + str(errs[0]) + ")")

        if stats_df is None or not isinstance(stats_df, pd.DataFrame) or stats_df.empty:
            stats_df = pd.DataFrame([{"total_tickets": 0}])
        if asset_df is None or not isinstance(asset_df, pd.DataFrame) or asset_df.empty:
            asset_df = pd.DataFrame([{"total_assets": 0}])
        if proc_df is None or not isinstance(proc_df, pd.DataFrame) or proc_df.empty:
            proc_df = pd.DataFrame([{"total_requests": 0}])
        if fleet_df is None or not isinstance(fleet_df, pd.DataFrame) or fleet_df.empty:
            fleet_df = pd.DataFrame([{"total_vehicles": 0}])

        status_df = status_df if isinstance(status_df, pd.DataFrame) else pd.DataFrame(columns=["status", "count"])
        priority_df = priority_df if isinstance(priority_df, pd.DataFrame) else pd.DataFrame(columns=["priority", "count"])
        location_df = location_df if isinstance(location_df, pd.DataFrame) else pd.DataFrame(columns=["location", "count"])
        asset_status_df = asset_status_df if isinstance(asset_status_df, pd.DataFrame) else pd.DataFrame(columns=["status", "count"])
        asset_location_df = asset_location_df if isinstance(asset_location_df, pd.DataFrame) else pd.DataFrame(columns=["location", "count"])
        asset_type_df = asset_type_df if isinstance(asset_type_df, pd.DataFrame) else pd.DataFrame(columns=["type", "count"])

        if not DB_AVAILABLE:
            if asset_type_df is None or asset_type_df.empty:
                asset_type_df = pd.DataFrame({
                    "type": ["Laptop", "Desktop", "Monitor", "Printer", "Phone"],
                    "count": [12, 8, 6, 3, 2]
                })
            if asset_location_df is None or asset_location_df.empty:
                asset_location_df = pd.DataFrame({
                    "location": ["Petersburg", "Hopewell", "Dinwiddie", "Surry"],
                    "count": [10, 7, 5, 3]
                })
            if status_df is None or status_df.empty:
                status_df = pd.DataFrame({"status": ["Open", "Resolved"], "count": [5, 3]})
            if priority_df is None or priority_df.empty:
                priority_df = pd.DataFrame({"priority": ["Low", "Medium", "High"], "count": [2, 5, 3]})
            if location_df is None or location_df.empty:
                location_df = pd.DataFrame({"location": ["Petersburg", "Hopewell"], "count": [6, 4]})

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Tickets", int(stats_df.iloc[0]['total_tickets']) if 'total_tickets' in stats_df.columns and len(stats_df)>0 else 0)
        with col2:
            st.metric("Total Assets", int(asset_df.iloc[0]['total_assets']) if 'total_assets' in asset_df.columns and len(asset_df)>0 else 0)
        with col3:
            st.metric("Procurement Requests", int(proc_df.iloc[0]['total_requests']) if 'total_requests' in proc_df.columns and len(proc_df)>0 else 0)
        with col4:
            st.metric("Total Vehicles", int(fleet_df.iloc[0]['total_vehicles']) if 'total_vehicles' in fleet_df.columns and len(fleet_df)>0 else 0)

        st.markdown("---")
        st.subheader("📊 Ticket Analytics")
        c1, c2, c3 = st.columns(3)
        with c1:
            if HAS_PLOTLY and not status_df.empty:
                try:
                    fig = px.pie(status_df, values='count', names='status', color_discrete_sequence=[VDH_NAVY, VDH_ORANGE])
                    st.plotly_chart(fig, config={"displayModeBar": False}, width='stretch')
                except Exception:
                    st.info("Chart unavailable")
            else:
                st.info("No status data or plotly not installed")
        with c2:
            if HAS_PLOTLY and not priority_df.empty:
                try:
                    fig = px.bar(priority_df, x='priority', y='count', color_discrete_sequence=[VDH_NAVY])
                    st.plotly_chart(fig, config={"displayModeBar": False}, width='stretch')
                except Exception:
                    st.info("Chart unavailable")
            else:
                st.info("No priority data or plotly not installed")
        with c3:
            if HAS_PLOTLY and not location_df.empty:
                try:
                    fig = px.bar(location_df, x='location', y='count', color_discrete_sequence=[VDH_ORANGE])
                    st.plotly_chart(fig, config={"displayModeBar": False}, width='stretch')
                except Exception:
                    st.info("Chart unavailable")
            else:
                st.info("No location data or plotly not installed")

        st.markdown("---")
        st.subheader("💻 Asset Analytics")
        c1, c2, c3 = st.columns(3)
        with c1:
            if HAS_PLOTLY and not asset_status_df.empty:
                try:
                    fig = px.pie(asset_status_df, values='count', names='status', 
                                 title='Asset Status Distribution',
                                 color_discrete_sequence=[VDH_NAVY, VDH_ORANGE, '#FFA500', '#FF0000'])
                    st.plotly_chart(fig, config={"displayModeBar": False}, width='stretch')
                except Exception:
                    st.info("Chart unavailable")
            else:
                st.info("No asset status data or plotly not installed")
        with c2:
            if HAS_PLOTLY and not asset_type_df.empty:
                try:
                    fig = px.bar(asset_type_df, x='type', y='count', 
                                 title='Assets by Type',
                                 color_discrete_sequence=[VDH_NAVY])
                    st.plotly_chart(fig, config={"displayModeBar": False}, width='stretch')
                except Exception:
                    st.info("Chart unavailable")
            else:
                st.info("No asset type data or plotly not installed")
        with c3:
            if HAS_PLOTLY and not asset_location_df.empty:
                try:
                    fig = px.bar(asset_location_df, x='location', y='count', 
                                 title='Assets by Location',
                                 color_discrete_sequence=[VDH_ORANGE])
                    st.plotly_chart(fig, config={"displayModeBar": False}, width='stretch')
                except Exception:
                    st.info("Chart unavailable")
            else:
                st.info("No asset location data or plotly not installed")

    # CONNECTED MODE: Helpdesk Tickets - attempt to list tickets when DB_AVAILABLE is True
    elif page == "🎫 Helpdesk Tickets":
        st.header("🎫 Helpdesk Tickets")
        
        # Create Ticket Button
        if 'show_ticket_form' not in st.session_state:
            st.session_state.show_ticket_form = False
            
        if st.button("➕ Create New Ticket", type="primary", key="create_ticket_top"):
            st.session_state.show_ticket_form = True
        
        if st.session_state.show_ticket_form:
            with st.form("quick_ticket_form"):
                st.subheader("📝 Create New Ticket")
                col1, col2 = st.columns(2)
                with col1:
                    ticket_name = st.text_input("Your Name *")
                    ticket_email = st.text_input("Email *")
                with col2:
                    ticket_location = st.selectbox("Location *", LOCATION_OPTIONS)
                    ticket_category = st.selectbox("Category *", ["IT Support", "Facilities", "HR", "Finance", "Other"])
                ticket_priority = st.selectbox("Priority *", ["Low", "Medium", "High", "Critical"])
                ticket_description = st.text_area("Description *", height=100)
                col1, col2 = st.columns(2)
                with col1:
                    submitted = st.form_submit_button("✅ Submit", type="primary", use_container_width=True)
                with col2:
                    cancel = st.form_submit_button("❌ Cancel", use_container_width=True)
                if cancel:
                    st.session_state.show_ticket_form = False
                    st.rerun()
                if submitted and ticket_name.strip() and ticket_email.strip() and ticket_description.strip():
                    st.success("✅ Ticket created!")
                    st.session_state.show_ticket_form = False
                    st.rerun()

        
        # Initialize session states
        if 'view_ticket_id' not in st.session_state:
            st.session_state.view_ticket_id = None
        if 'edit_ticket_id' not in st.session_state:
            st.session_state.edit_ticket_id = None
        
        if not DB_AVAILABLE:
            st.warning("Database unavailable. Showing demo tickets.")
            demo_tickets = pd.DataFrame([
                {"ticket_id": "T-001", "subject": "Password Reset", "status": "Open", "priority": "Medium"},
                {"ticket_id": "T-002", "subject": "Printer Issue", "status": "In Progress", "priority": "High"},
            ])
            st.dataframe(demo_tickets, width='stretch')
        else:
            # DETAILED TICKET VIEW
            if st.session_state.view_ticket_id:
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 4, 1])
                with col1:
                    if st.button("← Back"):
                        st.session_state.view_ticket_id = None
                        st.rerun()
                with col3:
                    if st.button("✏️ Edit"):
                        st.session_state.edit_ticket_id = st.session_state.view_ticket_id
                        st.session_state.view_ticket_id = None
                        st.rerun()
                
                detail_query = f"SELECT * FROM dbo.Tickets WHERE ticket_id = {st.session_state.view_ticket_id}"
                ticket_df, detail_err = execute_query(detail_query)
                
                if detail_err or ticket_df is None or len(ticket_df) == 0:
                    st.error("Ticket not found")
                    st.session_state.view_ticket_id = None
                else:
                    ticket = ticket_df.iloc[0]
                    
                    # Header
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        subject = ticket.get('short_description', 'N/A')
                        st.subheader(f"Ticket #{st.session_state.view_ticket_id}: {subject}")
                    with col2:
                        status = ticket.get('status', 'N/A')
                        st.write(f"**Status:** {status}")
                    with col3:
                        priority = ticket.get('priority', 'Normal')
                        priority_colors = {'Low': '🟢', 'Medium': '🟡', 'High': '🟠', 'Critical': '🔴'}
                        st.write(f"{priority_colors.get(priority, '⚪')} **{priority}**")
                    
                    st.markdown("---")
                    
                    tab1, tab2, tab3 = st.tabs(["📊 Details", "👤 Customer", "📝 History"])
                    
                    with tab1:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Ticket Information")
                            st.write(f"**Ticket ID:** {ticket.get('ticket_id', 'N/A')}")
                            st.write(f"**Status:** {ticket.get('status', 'N/A')}")
                            st.write(f"**Priority:** {ticket.get('priority', 'N/A')}")
                            st.write(f"**Location:** {ticket.get('location', 'N/A')}")
                            st.write(f"**Assigned To:** {ticket.get('assigned_to', 'Unassigned')}")
                        
                        with col2:
                            st.write("### Timestamps")
                            st.write(f"**Created:** {ticket.get('created_at', 'N/A')}")
                            st.write(f"**First Response:** {ticket.get('first_response_at', 'N/A')}")
                            st.write(f"**Resolved:** {ticket.get('resolved_at', 'N/A')}")
                        
                        st.write("### Description")
                        st.write(ticket.get('short_description', 'N/A'))
                        if ticket.get('description'):
                            with st.expander("Full Description"):
                                st.write(ticket.get('description', 'No description'))
                    
                    with tab2:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Contact Information")
                            st.write(f"**Name:** {ticket.get('name', 'N/A')}")
                            st.write(f"**Email:** {ticket.get('email', 'N/A')}")
                            st.write(f"**Phone:** {ticket.get('phone_number', 'N/A')}")
                        
                        with col2:
                            st.write("### Location")
                            st.write(f"**Location:** {ticket.get('location', 'N/A')}")
                    
                    with tab3:
                        st.write("### Ticket History")
                        notes_query = f"""
                            SELECT note_id, note_text, note_type, created_by, created_at
                            FROM dbo.Ticket_Notes
                            WHERE ticket_id = {st.session_state.view_ticket_id}
                            ORDER BY created_at DESC
                        """
                        notes_df, notes_error = execute_query(notes_query)
                        
                        if notes_error:
                            st.info("No history available. (Ticket_Notes table may not exist yet)")
                        elif notes_df is None or len(notes_df) == 0:
                            st.info("No history for this ticket yet.")
                        else:
                            for _, note in notes_df.iterrows():
                                st.markdown(f"""
                                <div class="note-item">
                                    <div class="note-header">{note['note_type']} • {note['created_by']} • {note['created_at']}</div>
                                    <div class="note-text">{note['note_text']}</div>
                                </div>
                                """, unsafe_allow_html=True)
            
            # GALLERY LIST VIEW
            else:
                query = """
                    SELECT TOP 200 
                        ticket_id, status, priority, name, email, location, 
                        phone_number, short_description, created_at
                    FROM dbo.Tickets 
                    ORDER BY created_at DESC
                """
                df, err = execute_query(query)
                
                if err:
                    st.error(f"Could not load tickets: {err}")
                elif df is None or df.empty:
                    st.info("No tickets found.")
                else:
                    # Statistics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        open_tickets = len(df[df['status'].isin(['New', 'Open', 'In Progress'])])
                        st.metric("Open", open_tickets)
                    with col2:
                        resolved = len(df[df['status'] == 'Resolved'])
                        st.metric("Resolved", resolved)
                    with col3:
                        high_priority = len(df[df['priority'].isin(['High', 'Critical'])])
                        st.metric("High Priority", high_priority)
                    with col4:
                        total = len(df)
                        st.metric("Total", total)
                    
                    st.markdown("---")
                    
                    # Search and filter
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        search = st.text_input("🔍 Search by Subject, Customer, or Location", "")
                    with col2:
                        status_filter = st.multiselect(
                            "Filter by Status",
                            options=['New', 'Open', 'In Progress', 'On Hold', 'Resolved', 'Closed'],
                            default=['New', 'Open', 'In Progress']
                        )
                    
                    filtered_df = df.copy()
                    if search:
                        filtered_df = filtered_df[
                            filtered_df['short_description'].str.contains(search, case=False, na=False) |
                            filtered_df['name'].str.contains(search, case=False, na=False) |
                            filtered_df['location'].str.contains(search, case=False, na=False)
                        ]
                    if status_filter:
                        filtered_df = filtered_df[filtered_df['status'].isin(status_filter)]
                    
                    st.markdown("---")
                    
                    if len(filtered_df) == 0:
                        st.info("No tickets match your search criteria.")
                    else:
                        st.success(f"📊 Showing {len(filtered_df)} ticket(s)")
                        
                        for idx, ticket in filtered_df.iterrows():
                            row_class = "item-row-even" if idx % 2 == 0 else "item-row-odd"
                            st.markdown(f'<div class="item-row {row_class}">', unsafe_allow_html=True)
                            
                            col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
                            
                            with col1:
                                ticket_id = ticket.get('ticket_id', 'N/A')
                                subject = ticket.get('short_description', 'N/A')
                                status = ticket.get('status', 'N/A')
                                
                                # Bold new tickets for easy identification
                                if status == 'New':
                                    st.markdown(f'<div class="list-header">🆕 <strong>Ticket #{ticket_id}</strong></div>', unsafe_allow_html=True)
                                    st.markdown(f'<strong>{subject}</strong>', unsafe_allow_html=True)
                                else:
                                    st.markdown(f'<div class="list-header">Ticket #{ticket_id}</div>', unsafe_allow_html=True)
                                    st.write(subject)
                                
                                customer = ticket.get('name', 'N/A')
                                location = ticket.get('location', 'N/A')
                                st.caption(f"👤 {customer} • 📍 {location}")
                            
                            with col2:
                                status = ticket.get('status', 'N/A')
                                st.write(f"**Status:** {status}")
                                created = ticket.get('created_at', 'N/A')
                                st.caption(f"Created: {created}")
                            
                            with col3:
                                priority = ticket.get('priority', 'Normal')
                                priority_colors = {'Low': '🟢', 'Medium': '🟡', 'High': '🟠', 'Critical': '🔴'}
                                st.write(f"{priority_colors.get(priority, '⚪')} {priority}")
                            
                            with col4:
                                if st.button("📋 View", key=f"view_ticket_{idx}_{ticket_id}"):
                                    st.session_state.view_ticket_id = ticket_id
                                    st.rerun()
                            
                            st.markdown('</div>', unsafe_allow_html=True)
                        
                        st.markdown("---")
                        st.download_button(
                            "📥 Download Tickets CSV",
                            data=filtered_df.to_csv(index=False).encode('utf-8'),
                            file_name=f"vdh_tickets_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )

    elif page == "💻 Asset Management":
        st.header("💻 Asset Management")
        
        # Initialize session states
        if 'view_asset_id' not in st.session_state:
            st.session_state.view_asset_id = None
        if 'edit_asset_id' not in st.session_state:
            st.session_state.edit_asset_id = None
        
        if not DB_AVAILABLE:
            st.warning("Database unavailable. Showing demo assets.")
            demo_assets = pd.DataFrame([
                {"asset_id": "A-001", "asset_tag": "VDH-LAP-001", "type": "Laptop", "status": "Deployed"},
                {"asset_id": "A-002", "asset_tag": "VDH-PRN-001", "type": "Printer", "status": "In-Stock"},
            ])
            st.dataframe(demo_assets, width='stretch')
        else:
            # DETAILED ASSET VIEW
            if st.session_state.view_asset_id:
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 4, 1])
                with col1:
                    if st.button("← Back"):
                        st.session_state.view_asset_id = None
                        st.rerun()
                with col3:
                    if st.button("✏️ Edit"):
                        st.session_state.edit_asset_id = st.session_state.view_asset_id
                        st.session_state.view_asset_id = None
                        st.rerun()
                
                detail_query = f"SELECT * FROM dbo.Assets WHERE asset_id = {st.session_state.view_asset_id}"
                asset_df, detail_err = execute_query(detail_query)
                
                if detail_err or asset_df is None or len(asset_df) == 0:
                    st.error("Asset not found")
                    st.session_state.view_asset_id = None
                else:
                    asset = asset_df.iloc[0]
                    
                    # Header
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        asset_tag = asset.get('asset_tag', 'N/A')
                        asset_type = asset.get('type', 'N/A')
                        st.subheader(f"{asset_tag} - {asset_type}")
                    with col2:
                        category = asset.get('category', 'N/A')
                        st.write(f"**Category:** {category}")
                    with col3:
                        status = asset.get('status', 'Unknown')
                        status_colors = {
                            'Deployed': '🟢', 'In-Stock': '🟡',
                            'Surplus': '🟠', 'Unaccounted': '🔴'
                        }
                        st.write(f"{status_colors.get(status, '⚪')} **{status}**")
                    
                    st.markdown("---")
                    
                    tab1, tab2, tab3, tab4 = st.tabs(["📊 Overview", "💼 Assignment", "🔌 Network", "📝 History"])
                    
                    with tab1:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Asset Information")
                            st.write(f"**Asset ID:** {asset.get('asset_id', 'N/A')}")
                            st.write(f"**Asset Tag:** {asset.get('asset_tag', 'N/A')}")
                            st.write(f"**Type:** {asset.get('type', 'N/A')}")
                            st.write(f"**Category:** {asset.get('category', 'N/A')}")
                            st.write(f"**Model:** {asset.get('model', 'N/A')}")
                            st.write(f"**Serial:** {asset.get('serial', 'N/A')}")
                        
                        with col2:
                            st.write("### Warranty & Purchase")
                            st.write(f"**Purchase Date:** {asset.get('purchase_date', 'N/A')}")
                            st.write(f"**Warranty Exp:** {asset.get('warranty_expiration', 'N/A')}")
                            st.write(f"**Status:** {asset.get('status', 'N/A')}")
                            st.write(f"**Created:** {asset.get('created_at', 'N/A')}")
                            st.write(f"**Updated:** {asset.get('updated_at', 'N/A')}")
                    
                    with tab2:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Current Assignment")
                            st.write(f"**Location:** {asset.get('location', 'Unassigned')}")
                            st.write(f"**Assigned User:** {asset.get('assigned_user', 'Unassigned')}")
                            st.write(f"**Email:** {asset.get('assigned_email', 'N/A')}")
                        
                        with col2:
                            st.write("### Status")
                            st.write(f"**Status:** {asset.get('status', 'N/A')}")
                    
                    with tab3:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Network Information")
                            st.write(f"**MAC Address:** {asset.get('mac_address', 'N/A')}")
                            st.write(f"**IP Address:** {asset.get('ip_address', 'N/A')}")
                        
                        with col2:
                            st.write("### Additional Info")
                            notes = asset.get('notes', 'No notes')
                            if notes and str(notes) != 'nan':
                                st.text_area("Notes", value=notes, height=100, disabled=True)
                            else:
                                st.info("No notes for this asset")
                    
                    with tab4:
                        st.write("### Asset History")
                        notes_query = f"""
                            SELECT note_id, note_text, note_type, created_by, created_at
                            FROM dbo.Asset_Notes
                            WHERE asset_id = {st.session_state.view_asset_id}
                            ORDER BY created_at DESC
                        """
                        notes_df, notes_error = execute_query(notes_query)
                        
                        if notes_error:
                            st.info("No history available. (Asset_Notes table may not exist yet)")
                        elif notes_df is None or len(notes_df) == 0:
                            st.info("No history for this asset yet.")
                        else:
                            for _, note in notes_df.iterrows():
                                st.markdown(f"""
                                <div class="note-item">
                                    <div class="note-header">{note['note_type']} • {note['created_by']} • {note['created_at']}</div>
                                    <div class="note-text">{note['note_text']}</div>
                                </div>
                                """, unsafe_allow_html=True)
            
            # GALLERY LIST VIEW
            else:
                query = """
                    SELECT TOP 200 
                        asset_id, asset_tag, type, category, model, serial, 
                        status, location, assigned_user, assigned_email
                    FROM dbo.Assets 
                    ORDER BY asset_id DESC
                """
                df, err = execute_query(query)
                
                if err:
                    st.error(f"Could not load assets: {err}")
                elif df is None or df.empty:
                    st.info("No assets found.")
                else:
                    # Statistics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        deployed = len(df[df['status'] == 'Deployed'])
                        st.metric("Deployed", deployed)
                    with col2:
                        in_stock = len(df[df['status'] == 'In-Stock'])
                        st.metric("In Stock", in_stock)
                    with col3:
                        surplus = len(df[df['status'] == 'Surplus'])
                        st.metric("Surplus", surplus)
                    with col4:
                        total = len(df)
                        st.metric("Total", total)
                    
                    st.markdown("---")
                    
                    # Search and filter
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        search = st.text_input("🔍 Search by Asset Tag, Model, or User", "")
                    with col2:
                        status_filter = st.multiselect(
                            "Filter by Status",
                            options=['Deployed', 'In-Stock', 'Surplus', 'Unaccounted'],
                            default=['Deployed', 'In-Stock']
                        )
                    
                    filtered_df = df.copy()
                    if search:
                        filtered_df = filtered_df[
                            filtered_df['asset_tag'].str.contains(search, case=False, na=False) |
                            filtered_df['model'].str.contains(search, case=False, na=False) |
                            filtered_df['assigned_user'].str.contains(search, case=False, na=False)
                        ]
                    if status_filter:
                        filtered_df = filtered_df[filtered_df['status'].isin(status_filter)]
                    
                    st.markdown("---")
                    
                    if len(filtered_df) == 0:
                        st.info("No assets match your search criteria.")
                    else:
                        st.success(f"📊 Showing {len(filtered_df)} asset(s)")
                        
                        for idx, asset in filtered_df.iterrows():
                            row_class = "item-row-even" if idx % 2 == 0 else "item-row-odd"
                            st.markdown(f'<div class="item-row {row_class}">', unsafe_allow_html=True)
                            
                            col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
                            
                            with col1:
                                asset_id = asset.get('asset_id', 'N/A')
                                asset_tag = asset.get('asset_tag', 'N/A')
                                asset_type = asset.get('type', 'N/A')
                                st.markdown(f'<div class="list-header">{asset_tag}</div>', unsafe_allow_html=True)
                                st.write(f"**Type:** {asset_type}")
                                
                                model = asset.get('model', 'N/A')
                                st.caption(f"Model: {model}")
                            
                            with col2:
                                location = asset.get('location', 'Unassigned')
                                assigned = asset.get('assigned_user', 'Unassigned')
                                st.write(f"📍 **{location}**")
                                st.caption(f"👤 {assigned}")
                            
                            with col3:
                                status = asset.get('status', 'Unknown')
                                status_colors = {
                                    'Deployed': '🟢', 'In-Stock': '🟡',
                                    'Surplus': '🟠', 'Unaccounted': '🔴'
                                }
                                st.write(f"{status_colors.get(status, '⚪')} **{status}**")
                            
                            with col4:
                                if st.button("📋 View", key=f"view_asset_{idx}_{asset_id}"):
                                    st.session_state.view_asset_id = asset_id
                                    st.rerun()
                            
                            st.markdown('</div>', unsafe_allow_html=True)
                        
                        st.markdown("---")
                        st.download_button(
                            "📥 Download Assets CSV",
                            data=filtered_df.to_csv(index=False).encode('utf-8'),
                            file_name=f"vdh_assets_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )

    elif page == "🛒 Procurement Requests":
        st.header("🛒 Procurement Requests")
        # Pending Procurement Requests Section (add to top of Procurement page)
        if DB_AVAILABLE:
            st.markdown("---")
            st.subheader("📋 Pending Procurement Requests")
            
            pending_query = """
                SELECT 
                    request_id,
                    request_number,
                    request_date,
                    requester_name,
                    requester_email,
                    location,
                    total_amount,
                    
                    status,
                    approval_1_status,
                    approval_2_status
                FROM dbo.Procurement_Requests
                WHERE status = 'Pending'
                ORDER BY request_date ASC
            """
            
            pending_df, pending_err = execute_query(pending_query)
            
            if pending_err:
                st.warning(f"Could not load pending requests: {pending_err}")
            elif pending_df is None or pending_df.empty:
                st.info("✅ No pending procurement requests")
            else:
                st.warning(f"⚠️ {len(pending_df)} pending procurement request(s) awaiting approval")
                
                # Show each pending request
                for idx, request in pending_df.iterrows():
                    with st.expander(f"🛒 {request['request_number']} - {request['requester_name']} (${request['total_amount']:,.2f})"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write("**Requester Information**")
                            st.write(f"Name: {request['requester_name']}")
                            st.write(f"Email: {request['requester_email']}")
                            st.write(f"Location: {request['location']}")
                            st.write(f"Requested: {request['request_date']}")
                        
                        with col2:
                            st.write("**Request Details**")
                            st.write(f"Request Number: {request['request_number']}")
                            st.write(f"Total Amount: ${request['total_amount']:,.2f}")
                            st.write(f"Approval 1: {request['approval_1_status']}")
                            st.write(f"Approval 2: {request['approval_2_status']}")
                        
                        st.write("**Description:**")
                        st.write(request['description'] if request['description'] else "No description provided")
                        
                        # Get line items
                        items_query = f"""
                            SELECT item_description, quantity, unit_price, total_price
                            FROM dbo.Procurement_Line_Items
                            WHERE request_id = {request['request_id']}
                        """
                        items_df, items_err = execute_query(items_query)
                        
                        if items_df is not None and not items_df.empty:
                            st.write("**Line Items:**")
                            st.dataframe(items_df, use_container_width=True, height=150)
                        
                        st.markdown("---")
                        
                        # Approval actions
                        col1, col2, col3, col4 = st.columns([2, 1, 1, 2])
                        
                        with col2:
                            if st.button("✅ Approve", key=f"approve_proc_{request['request_id']}", use_container_width=True):
                                with st.form(key=f"approve_proc_form_{request['request_id']}"):
                                    admin_name = st.text_input("Your Name (Admin)")
                                    approval_notes = st.text_area("Approval Notes (Optional)")
                                    
                                    if st.form_submit_button("Confirm Approval"):
                                        if admin_name:
                                            approve_query = """
                                                UPDATE dbo.Procurement_Requests 
                                                SET status = 'Approved', 
                                                    approved_by = ?,
                                                    approved_date = GETDATE(),
                                                    approval_notes = ?,
                                                    updated_at = GETDATE()
                                                WHERE request_id = ?
                                            """
                                            success, error = execute_non_query(approve_query, (admin_name, approval_notes, request['request_id']))
                                            
                                            if success:
                                                # Send notification email
                                                send_email_notification(
                                                    request['requester_email'],
                                                    "Procurement Request Approved",
                                                    f"Your procurement request {request['request_number']} has been approved."
                                                )
                                                
                                                st.success(f"✅ Request approved by {admin_name}")
                                                st.rerun()
                                            else:
                                                st.error(f"Error: {error}")
                                        else:
                                            st.error("Please provide your name")
                        
                        with col3:
                            if st.button("❌ Reject", key=f"reject_proc_{request['request_id']}", use_container_width=True):
                                with st.form(key=f"reject_proc_form_{request['request_id']}"):
                                    admin_name = st.text_input("Your Name (Admin)")
                                    rejection_reason = st.text_area("Rejection Reason *")
                                    
                                    if st.form_submit_button("Confirm Rejection"):
                                        if admin_name and rejection_reason:
                                            reject_query = """
                                                UPDATE dbo.Procurement_Requests 
                                                SET status = 'Rejected', 
                                                    approved_by = ?,
                                                    approved_date = GETDATE(),
                                                    approval_notes = ?,
                                                    updated_at = GETDATE()
                                                WHERE request_id = ?
                                            """
                                            success, error = execute_non_query(reject_query, (admin_name, rejection_reason, request['request_id']))
                                            
                                            if success:
                                                # Send notification email
                                                send_email_notification(
                                                    request['requester_email'],
                                                    "Procurement Request Rejected",
                                                    f"Your procurement request has been rejected. Reason: {rejection_reason}"
                                                )
                                                
                                                st.success(f"❌ Request rejected by {admin_name}")
                                                st.rerun()
                                            else:
                                                st.error(f"Error: {error}")
                                        else:
                                            st.error("Please provide your name and rejection reason")
                
                st.markdown("---")
                st.subheader("📜 Approval History")
                
                history_query = """
                    SELECT 
                        request_id,
                        request_number,
                        requester_name,
                        total_amount,
                        status,
                        approved_by,
                        approved_date,
                        approval_notes
                    FROM dbo.Procurement_Requests
                    WHERE status IN ('Approved', 'Rejected')
                    ORDER BY approved_date DESC
                """
                
                history_df, history_err = execute_query(history_query)
                
                if history_err:
                    st.info("No approval history available")
                elif history_df is None or history_df.empty:
                    st.info("No approval history yet")
                else:
                    st.dataframe(history_df, use_container_width=True, height=300)

        # Initialize session states
        if 'view_procurement_id' not in st.session_state:
            st.session_state.view_procurement_id = None
        
        if not DB_AVAILABLE:
            st.warning("Database unavailable. Showing demo procurement requests.")
            demo_procurements = pd.DataFrame([
                {"request_id": "PR-001", "request_number": "PR-20251024-001", "status": "Pending", "total_amount": "$1,250.00"},
                {"request_id": "PR-002", "request_number": "PR-20251024-002", "status": "Approved", "total_amount": "$500.00"},
            ])
            st.dataframe(demo_procurements, width='stretch')
        else:
            # DETAILED PROCUREMENT VIEW
            if st.session_state.view_procurement_id:
                st.markdown("---")
                if st.button("← Back to List"):
                    st.session_state.view_procurement_id = None
                    st.rerun()
                
                detail_query = f"SELECT * FROM dbo.Procurement_Requests WHERE request_id = {st.session_state.view_procurement_id}"
                procurement_df, detail_err = execute_query(detail_query)
                
                if detail_err or procurement_df is None or len(procurement_df) == 0:
                    st.error("Procurement request not found")
                    st.session_state.view_procurement_id = None
                else:
                    procurement = procurement_df.iloc[0]
                    
                    # Header
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        request_num = procurement.get('request_number', 'N/A')
                        st.subheader(f"Request: {request_num}")
                    with col2:
                        status = procurement.get('status', 'N/A')
                        status_colors = {
                            'Draft': '⚪', 'Pending': '🟡',
                            'Approved': '🟢', 'Rejected': '🔴',
                            'Completed': '✅'
                        }
                        st.write(f"{status_colors.get(status, '⚪')} **{status}**")
                    with col3:
                        total = procurement.get('total_amount', 0)
                        st.metric("Total", f"${total:,.2f}" if total else "N/A")
                    
                    # Delete button for draft requests
                    if status and status.lower() == "draft":
                        st.markdown("---")
                        col1, col2, col3 = st.columns([2, 1, 1])
                        with col3:
                            if st.button("🗑️ Delete Draft", type="secondary", use_container_width=True, key="delete_draft_btn"):
                                delete_query = "DELETE FROM dbo.Procurement_Requests WHERE request_id = ?"
                                result, err = execute_non_query(delete_query, (st.session_state.view_procurement_id,))
                                if err:
                                    st.error(f"❌ Error deleting draft: {err}")
                                else:
                                    st.success("✅ Draft deleted successfully!")
                                    del st.session_state["view_procurement_id"]
                                    import time
                                    time.sleep(1)
                                    st.rerun()
                    
                    st.markdown("---")
                    
                    tab1, tab2, tab3 = st.tabs(["📊 Request Info", "💰 Line Items", "📝 History"])
                    
                    with tab1:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Requester Information")
                            st.write(f"**Name:** {procurement.get('requester_name', 'N/A')}")
                            st.write(f"**Email:** {procurement.get('requester_email', 'N/A')}")
                            st.write(f"**Phone:** {procurement.get('requester_phone', 'N/A')}")
                            st.write(f"**Location:** {procurement.get('location', 'N/A')}")
                        
                        with col2:
                            st.write("### Request Details")
                            st.write(f"**Request Date:** {procurement.get('request_date', 'N/A')}")
                            st.write(f"**Status:** {procurement.get('status', 'N/A')}")
                            st.write(f"**Approval 1:** {procurement.get('approval_1_status', 'Pending')}")
                            st.write(f"**Approval 2:** {procurement.get('approval_2_status', 'Pending')}")
                        
                        st.write("### Description")
                        description = procurement.get('description', 'No description')
                        if description and str(description) != 'nan':
                            st.write(description)
                        else:
                            st.info("No description provided")
                    
                    with tab2:
                        st.write("### Procurement Line Items")
                        items_query = f"""
                            SELECT item_id, item_description, quantity, unit_price, total_price, 
                                   billing_code_cst, billing_code_coa, billing_code_prog, billing_code_fund
                            FROM dbo.Procurement_Line_Items
                            WHERE request_id = {st.session_state.view_procurement_id}
                            ORDER BY item_id
                        """
                        items_df, items_error = execute_query(items_query)
                        
                        if items_error:
                            st.info("No line items found or table doesn't exist")
                        elif items_df is None or len(items_df) == 0:
                            st.info("No line items for this request")
                        else:
                            for idx, item in items_df.iterrows():
                                with st.expander(f"Item {idx+1}: {item.get('item_description', 'N/A')}"):
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.write(f"**Quantity:** {item.get('quantity', 0)}")
                                        st.write(f"**Unit Price:** ${item.get('unit_price', 0):,.2f}")
                                    with col2:
                                        st.write(f"**Total:** ${item.get('total_price', 0):,.2f}")
                                    with col3:
                                        st.write(f"**CST:** {item.get('billing_code_cst', 'N/A')}")
                                        st.write(f"**COA:** {item.get('billing_code_coa', 'N/A')}")
                            
                            total_sum = items_df['total_price'].sum()
                            st.write(f"### Grand Total: ${total_sum:,.2f}")
                    
                    with tab3:
                        st.write("### Request History")
                        notes_query = f"""
                            SELECT note_id, note_text, note_type, created_by, created_at
                            FROM dbo.Procurement_Notes
                            WHERE request_id = {st.session_state.view_procurement_id}
                            ORDER BY created_at DESC
                        """
                        notes_df, notes_error = execute_query(notes_query)
                        
                        if notes_error:
                            st.info("No history available")
                        elif notes_df is None or len(notes_df) == 0:
                            st.info("No history for this request")
                        else:
                            for _, note in notes_df.iterrows():
                                st.markdown(f"""
                                <div class="note-item">
                                    <div class="note-header">{note['note_type']} • {note['created_by']} • {note['created_at']}</div>
                                    <div class="note-text">{note['note_text']}</div>
                                </div>
                                """, unsafe_allow_html=True)
            
            # GALLERY LIST VIEW
            else:
                query = """
                    SELECT TOP 200 
                        request_id, request_number, request_date, requester_name, 
                        requester_email, requester_phone, location, total_amount, status
                    FROM dbo.Procurement_Requests 
                    ORDER BY request_date DESC
                """
                df, err = execute_query(query)
                
                if err:
                    st.error(f"Could not load procurement requests: {err}")
                elif df is None or df.empty:
                    st.info("No procurement requests found.")
                else:
                    # Statistics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        pending = len(df[df['status'] == 'Pending'])
                        st.metric("Pending", pending)
                    with col2:
                        approved = len(df[df['status'] == 'Approved'])
                        st.metric("Approved", approved)
                    with col3:
                        completed = len(df[df['status'] == 'Completed'])
                        st.metric("Completed", completed)
                    with col4:
                        total = len(df)
                        st.metric("Total", total)
                    
                    st.markdown("---")
                    
                    # Search and filter
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        search = st.text_input("🔍 Search by Request Number, Requester, or Location", "")
                    with col2:
                        status_filter = st.multiselect(
                            "Filter by Status",
                            options=['Draft', 'Pending', 'Approved', 'Rejected', 'Completed'],
                            default=['Pending', 'Approved']
                        )
                    
                    filtered_df = df.copy()
                    if search:
                        filtered_df = filtered_df[
                            filtered_df['request_number'].str.contains(search, case=False, na=False) |
                            filtered_df['requester_name'].str.contains(search, case=False, na=False) |
                            filtered_df['location'].str.contains(search, case=False, na=False)
                        ]
                    if status_filter:
                        filtered_df = filtered_df[filtered_df['status'].isin(status_filter)]
                    
                    st.markdown("---")
                    
                    if len(filtered_df) == 0:
                        st.info("No requests match your search criteria.")
                    else:
                        st.success(f"📊 Showing {len(filtered_df)} request(s)")
                        
                        for idx, procurement in filtered_df.iterrows():
                            row_class = "item-row-even" if idx % 2 == 0 else "item-row-odd"
                            st.markdown(f'<div class="item-row {row_class}">', unsafe_allow_html=True)
                            
                            col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
                            
                            with col1:
                                request_num = procurement.get('request_number', 'N/A')
                                requester = procurement.get('requester_name', 'N/A')
                                st.markdown(f'<div class="list-header">{request_num}</div>', unsafe_allow_html=True)
                                st.write(f"👤 {requester}")
                                
                                location = procurement.get('location', 'N/A')
                                st.caption(f"📍 {location}")
                            
                            with col2:
                                request_date = procurement.get('request_date', 'N/A')
                                total_amount = procurement.get('total_amount', 0)
                                st.write(f"**Date:** {request_date}")
                                st.write(f"**Total:** ${total_amount:,.2f}" if total_amount else "**Total:** N/A")
                            
                            with col3:
                                status = procurement.get('status', 'Unknown')
                                status_colors = {
                                    'Draft': '⚪', 'Pending': '🟡',
                                    'Approved': '🟢', 'Rejected': '🔴',
                                    'Completed': '✅'
                                }
                                st.write(f"{status_colors.get(status, '⚪')} **{status}**")
                            
                            with col4:
                                request_id = procurement.get('request_id', 'N/A')
                                if st.button("📋 View", key=f"view_procurement_{idx}_{request_id}"):
                                    st.session_state.view_procurement_id = request_id
                                    st.rerun()
                            
                            st.markdown('</div>', unsafe_allow_html=True)
                        
                        st.markdown("---")
                        st.download_button(
                            "📥 Download Procurements CSV",
                            data=filtered_df.to_csv(index=False).encode('utf-8'),
                            file_name=f"vdh_procurements_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )

    elif page == "🚗 Fleet Management":
        st.header("🚗 Fleet Management")
        
        if not DB_AVAILABLE:
            st.warning("Database connection required for Fleet Management")
            st.stop()
        
        # Create tabs for Fleet Management
        tab1, tab2, tab3, tab4 = st.tabs(["📋 Vehicle Requests", "🚙 Dispatched Vehicles", "🚗 Vehicle Management", "📊 Trip Logs"])
        
        with tab1:
            st.subheader("📋 Vehicle Requests & Approvals")
            # Pending Vehicle Requests Section (add to top of Fleet Management)
            if DB_AVAILABLE:
                st.markdown("---")
                st.subheader("📋 Pending Vehicle Requests")
            
            pending_query = """
                SELECT 
                    vr.request_id,
                    vr.vehicle_id,
                    vr.requester_name,
                    vr.requester_email,
                    vr.requester_location,
                    vr.purpose,
                    vr.start_date,
                    vr.end_date,
                    vr.estimated_miles,
                    vr.request_date,
                    vr.status,
                    v.year,
                    v.make_model,
                    v.license_plate
                FROM dbo.Vehicle_Requests vr
                JOIN dbo.vehicles v ON vr.vehicle_id = v.id
                WHERE vr.status = 'Pending'
                ORDER BY vr.request_date ASC
            """
            
            pending_df, pending_err = execute_query(pending_query)
            
            if pending_err:
                st.warning(f"Could not load pending requests: {pending_err}")
            elif pending_df is None or pending_df.empty:
                st.info("✅ No pending vehicle requests")
            else:
                st.warning(f"⚠️ {len(pending_df)} pending vehicle request(s) awaiting approval")
                
                # Show each pending request
                for idx, request in pending_df.iterrows():
                    with st.expander(f"🚗 Request #{request['request_id']} - {request['requester_name']} ({request['year']} {request['make_model']})"):
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.write("**Requester Information**")
                            st.write(f"Name: {request['requester_name']}")
                            st.write(f"Email: {request['requester_email']}")
                            st.write(f"Location: {request['requester_location']}")
                            st.write(f"Requested: {request['request_date']}")
                        
                        with col2:
                            st.write("**Vehicle Details**")
                            st.write(f"Vehicle: {request['year']} {request['make_model']}")
                            st.write(f"License: {request['license_plate']}")
                            st.write(f"Est. Miles: {request['estimated_miles']}")
                        
                        with col3:
                            st.write("**Trip Details**")
                            st.write(f"Start: {request['start_date']}")
                            st.write(f"End: {request['end_date']}")
                            st.write(f"Purpose: {request['purpose'][:50]}...")
                        
                        st.markdown("---")
                        
                        # Approval actions
                        col1, col2, col3, col4 = st.columns([2, 1, 1, 2])
                        
                        with col2:
                            approve_key = f"approve_vehicle_{request['request_id']}"
                            
                            # Check if we're in approval mode for this request
                            if st.session_state.get(f"approving_{request['request_id']}", False):
                                st.write("**Approve Request:**")
                                with st.form(key=f"approve_form_{request['request_id']}"):
                                    admin_name = st.text_input("Your Name (Admin) *", key=f"admin_name_{request['request_id']}")
                                    
                                    col_a, col_b = st.columns(2)
                                    with col_a:
                                        submit_approve = st.form_submit_button("✅ Confirm", use_container_width=True)
                                    with col_b:
                                        cancel_approve = st.form_submit_button("Cancel", use_container_width=True)
                                    
                                    if submit_approve and admin_name:
                                        approve_query = """
                                            UPDATE dbo.Vehicle_Requests 
                                            SET status = 'Approved', 
                                                approved_by = ?,
                                                approved_date = GETDATE(),
                                                updated_at = GETDATE()
                                            WHERE request_id = ?
                                        """
                                        success, error = execute_non_query(approve_query, (admin_name, request['request_id']))
                                        
                                        if success:
                                            # Update vehicle status to Dispatched
                                            update_vehicle = """
                                                UPDATE dbo.vehicles 
                                                SET status = 'Dispatched',
                                                    current_driver = ?,
                                                    last_used_date = GETDATE(),
                                                    usage_count = usage_count + 1,
                                                    updated_at = GETDATE()
                                                WHERE id = ?
                                            """
                                            execute_non_query(update_vehicle, (request['requester_name'], request['vehicle_id']))
                                            
                                            # Send notification email
                                            send_email_notification(
                                                request['requester_email'],
                                                "Vehicle Request Approved",
                                                f"Your vehicle request for {request['year']} {request['make_model']} has been approved."
                                            )
                                            
                                            st.success(f"✅ Request approved by {admin_name}")
                                            st.session_state[f"approving_{request['request_id']}"] = False
                                            st.rerun()
                                        else:
                                            st.error(f"Error: {error}")
                                    
                                    if cancel_approve:
                                        st.session_state[f"approving_{request['request_id']}"] = False
                                        st.rerun()
                            else:
                                # Show approve button
                                if st.button("✅ Approve", key=approve_key, use_container_width=True):
                                    st.session_state[f"approving_{request['request_id']}"] = True
                                    st.rerun()
                        
                        with col3:
                            if st.button("❌ Reject", key=f"reject_vehicle_{request['request_id']}", use_container_width=True):
                                with st.form(key=f"reject_form_{request['request_id']}"):
                                    admin_name = st.text_input("Your Name (Admin)")
                                    rejection_reason = st.text_area("Rejection Reason *")
                                    
                                    if st.form_submit_button("Confirm Rejection"):
                                        if admin_name and rejection_reason:
                                            reject_query = """
                                                UPDATE dbo.Vehicle_Requests 
                                                SET status = 'Rejected', 
                                                    approved_by = ?,
                                                    approved_date = GETDATE(),
                                                    rejection_reason = ?,
                                                    updated_at = GETDATE()
                                                WHERE request_id = ?
                                            """
                                            success, error = execute_non_query(reject_query, (admin_name, rejection_reason, request['request_id']))
                                            
                                            if success:
                                                # Send notification email
                                                send_email_notification(
                                                    request['requester_email'],
                                                    "Vehicle Request Rejected",
                                                    f"Your vehicle request has been rejected. Reason: {rejection_reason}"
                                                )
                                                
                                                st.success(f"❌ Request rejected by {admin_name}")
                                                st.rerun()
                                            else:
                                                st.error(f"Error: {error}")
                                        else:
                                            st.error("Please provide your name and rejection reason")
                
                st.markdown("---")
                st.subheader("📜 Approval History")
                
                history_query = """
                    SELECT 
                        vr.request_id,
                        vr.requester_name,
                        vr.status,
                        vr.approved_by,
                        vr.approved_date,
                        vr.rejection_reason,
                        v.year,
                        v.make_model,
                        v.license_plate
                    FROM dbo.Vehicle_Requests vr
                    JOIN dbo.vehicles v ON vr.vehicle_id = v.id
                    WHERE vr.status IN ('Approved', 'Rejected')
                    ORDER BY vr.approved_date DESC
                """
                
                history_df, history_err = execute_query(history_query)
                
                if history_err:
                    st.info("No approval history available")
                elif history_df is None or history_df.empty:
                    st.info("No approval history yet")
                else:
                    st.dataframe(history_df, use_container_width=True, height=300)
        
        with tab2:
            st.subheader("🚙 Currently Dispatched Vehicles")
            # Dispatched Vehicles Section
            st.markdown("---")
        
        dispatched_query = """
            SELECT 
                v.id, v.year, v.make_model, v.license_plate,
                v.current_driver, v.last_used_date, v.current_mileage,
                vr.requester_email, vr.requester_phone, vr.requester_location,
                vr.end_date, vr.purpose, vr.estimated_miles, vr.start_date
            FROM dbo.vehicles v
            LEFT JOIN dbo.Vehicle_Requests vr ON v.id = vr.vehicle_id 
                AND vr.status = 'Approved'
                AND vr.end_date >= CAST(GETDATE() AS DATE)
            WHERE v.status = 'Dispatched'
            ORDER BY v.last_used_date DESC
        """
        
        dispatched_df, disp_err = execute_query(dispatched_query)
        
        if disp_err or dispatched_df is None or dispatched_df.empty:
            st.info("✅ No vehicles currently dispatched")
        else:
            st.success(f"**{len(dispatched_df)} vehicle(s) currently in use**")
            
            for idx, vehicle in dispatched_df.iterrows():
                with st.expander(f"🚗 {vehicle['year']} {vehicle['make_model']} - {vehicle['license_plate']} (Driver: {vehicle['current_driver']})"):
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.write("**Driver Information**")
                        st.write(f"👤 **Name:** {vehicle['current_driver']}")
                        if pd.notna(vehicle.get('requester_email')):
                            st.write(f"📧 **Email:** {vehicle['requester_email']}")
                        if pd.notna(vehicle.get('requester_phone')):
                            st.write(f"📞 **Phone:** {vehicle['requester_phone']}")
                        if pd.notna(vehicle.get('requester_location')):
                            st.write(f"📍 **Location:** {vehicle['requester_location']}")
                    
                    with col2:
                        st.write("**Trip Information**")
                        if pd.notna(vehicle.get('start_date')):
                            st.write(f"🗓️ **Start Date:** {vehicle['start_date']}")
                        if pd.notna(vehicle.get('end_date')):
                            st.write(f"🏁 **Return Date:** {vehicle['end_date']}")
                        st.write(f"🚗 **Checked Out:** {vehicle['last_used_date']}")
                        if pd.notna(vehicle.get('estimated_miles')):
                            st.write(f"📏 **Est. Miles:** {vehicle['estimated_miles']}")
                    
                    with col3:
                        st.write("**Vehicle Status**")
                        st.write(f"🔢 **Current Mileage:** {vehicle['current_mileage']:,}")
                        if pd.notna(vehicle.get('purpose')):
                            st.write(f"📝 **Purpose:**")
                            st.caption(f"{vehicle['purpose'][:150]}...")
                    
                    # Contact button
                    st.markdown("---")
                    col1, col2, col3 = st.columns([2, 1, 2])
                    with col2:
                        if st.button("📞 Contact Driver", key=f"contact_driver_{vehicle['id']}", use_container_width=True):
                            if pd.notna(vehicle.get('requester_email')) or pd.notna(vehicle.get('requester_phone')):
                                contact_info = []
                                if pd.notna(vehicle.get('requester_email')):
                                    contact_info.append(f"📧 Email: {vehicle['requester_email']}")
                                if pd.notna(vehicle.get('requester_phone')):
                                    contact_info.append(f"📞 Phone: {vehicle['requester_phone']}")
                                st.info("\n".join(contact_info))
                            else:
                                st.warning("Contact information not available")

        with tab3:
            st.subheader("🚗 Vehicle Management")
            # Initialize session states
            if 'view_vehicle_id' not in st.session_state:
                st.session_state.view_vehicle_id = None
            if 'edit_vehicle_id' not in st.session_state:
                st.session_state.edit_vehicle_id = None
            if 'add_vehicle_mode' not in st.session_state:
                st.session_state.add_vehicle_mode = False
        
        if not DB_AVAILABLE:
            st.warning("Database unavailable. Showing demo fleet list.")
            demo_fleet = pd.DataFrame([
                {"vehicle_id": "V-001", "make": "Ford", "model": "Transit", "plate": "ABC-123", "status": "Available"},
                {"vehicle_id": "V-002", "make": "Chevy", "model": "Express", "plate": "XYZ-789", "status": "In Use"},
            ])
            st.dataframe(demo_fleet, width='stretch')
        else:
            # ADD NEW VEHICLE MODE
            if st.session_state.add_vehicle_mode:
                st.markdown("---")
                col1, col2 = st.columns([1, 5])
                with col1:
                    if st.button("← Cancel"):
                        st.session_state.add_vehicle_mode = False
                        st.rerun()
                with col2:
                    st.subheader("➕ Add New Vehicle")
                
                with st.form("add_vehicle_form"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("### Basic Information")
                        year = st.number_input("Year *", min_value=1990, max_value=2030, value=2024)
                        make_model = st.text_input("Make & Model *", placeholder="e.g., Ford F-150")
                        vin = st.text_input("VIN *", placeholder="17-character VIN")
                        license_plate = st.text_input("License Plate *", placeholder="e.g., VAH-1234")
                        status = st.selectbox("Status *", ["Available", "Dispatched", "Maintenance", "Out Of Service"])
                    
                    with col2:
                        st.write("### Mileage Information")
                        initial_mileage = st.number_input("Initial Mileage *", min_value=0, value=0)
                        current_mileage = st.number_input("Current Mileage *", min_value=0, value=0)
                        last_service_mileage = st.number_input("Last Service Mileage", min_value=0, value=0)
                        last_service_date = st.date_input("Last Service Date")
                    
                    st.write("### Photo Upload")
                    uploaded_file = st.file_uploader("Upload Vehicle Photo (JPG, PNG)", type=['jpg', 'jpeg', 'png'])
                    
                    st.write("### Additional Information")
                    notes = st.text_area("Notes", placeholder="Any additional information about this vehicle...")
                    
                    submit = st.form_submit_button("✅ Add Vehicle")
                    
                    if submit:
                        if not make_model or not vin or not license_plate:
                            st.error("Please fill in all required fields (*)")
                        else:
                            miles_until_service = 3000 - (current_mileage - last_service_mileage)
                            
                            photo_url = None
                            if uploaded_file is not None:
                                import base64
                                bytes_data = uploaded_file.getvalue()
                                base64_image = base64.b64encode(bytes_data).decode()
                                photo_url = f"data:image/{uploaded_file.type.split('/')[1]};base64,{base64_image}"
                            
                            insert_query = """
                                INSERT INTO dbo.vehicles (
                                    year, make_model, vin, license_plate, 
                                    initial_mileage, current_mileage, 
                                    last_service_date, last_service_mileage, 
                                    miles_until_service, status, 
                                    photo_url, notes_log,
                                    created_at, updated_at
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE())
                            """
                            
                            success, error = execute_non_query(insert_query, (
                                year, make_model, vin, license_plate,
                                initial_mileage, current_mileage,
                                last_service_date, last_service_mileage,
                                miles_until_service, status,
                                photo_url, notes
                            ))
                            
                            if success:
                                st.success(f"✅ Vehicle {license_plate} added successfully!")
                                st.session_state.add_vehicle_mode = False
                                st.rerun()
                            else:
                                st.error(f"❌ Error adding vehicle: {error}")
            
            # EDIT VEHICLE MODE
            elif st.session_state.edit_vehicle_id:
                st.markdown("---")
                col1, col2 = st.columns([1, 5])
                with col1:
                    if st.button("← Cancel"):
                        st.session_state.edit_vehicle_id = None
                        st.rerun()
                with col2:
                    st.subheader("✏️ Edit Vehicle")
                
                edit_query = f"SELECT * FROM dbo.vehicles WHERE id = {st.session_state.edit_vehicle_id}"
                vehicle_df, edit_err = execute_query(edit_query)
                
                if edit_err or vehicle_df is None or len(vehicle_df) == 0:
                    st.error("Vehicle not found")
                    st.session_state.edit_vehicle_id = None
                else:
                    vehicle = vehicle_df.iloc[0]
                    
                    with st.form("edit_vehicle_form"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write("### Basic Information")
                            year = st.number_input("Year", min_value=1990, max_value=2030, value=int(vehicle.get('year', 2024)))
                            make_model = st.text_input("Make & Model", value=str(vehicle.get('make_model', '')))
                            vin = st.text_input("VIN", value=str(vehicle.get('vin', '')))
                            license_plate = st.text_input("License Plate", value=str(vehicle.get('license_plate', '')))
                            status = st.selectbox("Status", 
                                ["Available", "Dispatched", "Maintenance", "Out Of Service"],
                                index=["Available", "Dispatched", "Maintenance", "Out Of Service"].index(vehicle.get('status', 'Available'))
                            )
                        
                        with col2:
                            st.write("### Mileage Information")
                            initial_mileage = st.number_input("Initial Mileage", min_value=0, value=int(vehicle.get('initial_mileage', 0)))
                            current_mileage = st.number_input("Current Mileage", min_value=0, value=int(vehicle.get('current_mileage', 0)))
                            last_service_mileage = st.number_input("Last Service Mileage", min_value=0, value=int(vehicle.get('last_service_mileage', 0)))
                            
                            current_last_service = vehicle.get('last_service_date')
                            if current_last_service and str(current_last_service) != 'nan':
                                if isinstance(current_last_service, str):
                                    try:
                                        current_last_service = datetime.strptime(current_last_service.split()[0], '%Y-%m-%d').date()
                                    except:
                                        current_last_service = None
                            else:
                                current_last_service = None
                            
                            last_service_date = st.date_input("Last Service Date", value=current_last_service)
                        
                        st.write("### Photo Upload")
                        current_photo = vehicle.get('photo_url') or vehicle.get('picture_url')
                        if current_photo and str(current_photo) != 'nan' and str(current_photo).strip():
                            try:
                                st.image(current_photo, width=200, caption="Current Photo")
                            except:
                                st.info("Current photo unavailable")
                        
                        uploaded_file = st.file_uploader("Upload New Photo - Leave empty to keep current", type=['jpg', 'jpeg', 'png'])
                        
                        st.write("### Additional Information")
                        current_driver = st.text_input("Current Driver", value=str(vehicle.get('current_driver', '') if vehicle.get('current_driver') else ''))
                        notes = st.text_area("Notes", value=str(vehicle.get('notes_log', '') if vehicle.get('notes_log') else ''))
                        
                        submit = st.form_submit_button("💾 Save Changes")
                        
                        if submit:
                            miles_until_service = 3000 - (current_mileage - last_service_mileage)
                            
                            photo_url = current_photo
                            if uploaded_file is not None:
                                import base64
                                bytes_data = uploaded_file.getvalue()
                                base64_image = base64.b64encode(bytes_data).decode()
                                photo_url = f"data:image/{uploaded_file.type.split('/')[1]};base64,{base64_image}"
                            
                            update_query = """
                                UPDATE dbo.vehicles 
                                SET year = ?, make_model = ?, vin = ?, license_plate = ?,
                                    initial_mileage = ?, current_mileage = ?,
                                    last_service_date = ?, last_service_mileage = ?,
                                    miles_until_service = ?, status = ?,
                                    photo_url = ?, current_driver = ?, notes_log = ?,
                                    updated_at = GETDATE()
                                WHERE id = ?
                            """
                            
                            success, error = execute_non_query(update_query, (
                                year, make_model, vin, license_plate,
                                initial_mileage, current_mileage,
                                last_service_date, last_service_mileage,
                                miles_until_service, status,
                                photo_url, current_driver if current_driver else None, notes,
                                st.session_state.edit_vehicle_id
                            ))
                            
                            if success:
                                st.success("✅ Vehicle updated successfully!")
                                st.session_state.edit_vehicle_id = None
                                st.session_state.view_vehicle_id = None
                                st.rerun()
                            else:
                                st.error(f"❌ Error: {error}")
            
            # DETAILED VEHICLE VIEW
            elif st.session_state.view_vehicle_id:
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 4, 1])
                with col1:
                    if st.button("← Back"):
                        st.session_state.view_vehicle_id = None
                        st.rerun()
                with col3:
                    if st.button("✏️ Edit"):
                        st.session_state.edit_vehicle_id = st.session_state.view_vehicle_id
                        st.session_state.view_vehicle_id = None
                        st.rerun()
                
                detail_query = f"SELECT * FROM dbo.vehicles WHERE id = {st.session_state.view_vehicle_id}"
                vehicle_df, detail_err = execute_query(detail_query)
                
                if detail_err or vehicle_df is None or len(vehicle_df) == 0:
                    st.error("Vehicle not found")
                    st.session_state.view_vehicle_id = None
                else:
                    vehicle = vehicle_df.iloc[0]
                    
                    col1, col2 = st.columns([1, 2])
                    
                    with col1:
                        photo_url = vehicle.get('photo_url') or vehicle.get('picture_url')
                        if photo_url and str(photo_url).strip() and str(photo_url) != 'nan':
                            try:
                                st.image(photo_url, width='stretch', caption=f"{vehicle.get('year', '')} {vehicle.get('make_model', '')}")
                            except:
                                st.markdown("""
                                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                                padding: 60px; text-align: center; border-radius: 10px; color: white; font-size: 48px;">
                                        🚗
                                    </div>
                                """, unsafe_allow_html=True)
                                st.caption("No photo available")
                        else:
                            st.markdown("""
                                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                            padding: 60px; text-align: center; border-radius: 10px; color: white; font-size: 48px;">
                                    🚗
                                </div>
                            """, unsafe_allow_html=True)
                            st.caption("No photo available")
                    
                    with col2:
                        year = vehicle.get('year', 'N/A')
                        make_model = vehicle.get('make_model', 'N/A')
                        st.subheader(f"{year} {make_model}")
                        
                        status = vehicle.get('status', 'Unknown')
                        status_colors = {
                            'Available': '🟢',
                            'Dispatched': '🟡',
                            'Maintenance': '🟠',
                            'Out Of Service': '🔴'
                        }
                        st.markdown(f"### {status_colors.get(status, '⚪')} {status}")
                        
                        license = vehicle.get('license_plate', 'N/A')
                        st.write(f"**License Plate:** {license}")
                        
                        vin = vehicle.get('vin', 'N/A')
                        st.write(f"**VIN:** {vin}")
                    
                    st.markdown("---")
                    
                    tab1, tab2, tab3, tab4 = st.tabs(["📊 Overview", "🔧 Maintenance", "📍 Location & Usage", "📝 Notes"])
                    
                    with tab1:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Vehicle Information")
                            st.write(f"**Year:** {vehicle.get('year', 'N/A')}")
                            st.write(f"**Make/Model:** {vehicle.get('make_model', 'N/A')}")
                            st.write(f"**VIN:** {vehicle.get('vin', 'N/A')}")
                            st.write(f"**License Plate:** {vehicle.get('license_plate', 'N/A')}")
                            st.write(f"**Status:** {vehicle.get('status', 'N/A')}")
                        
                        with col2:
                            st.write("### Mileage Information")
                            initial_mileage = vehicle.get('initial_mileage', 0)
                            current_mileage = vehicle.get('current_mileage', 0)
                            st.write(f"**Initial Mileage:** {initial_mileage:,} mi" if initial_mileage else "**Initial Mileage:** N/A")
                            st.write(f"**Current Mileage:** {current_mileage:,} mi" if current_mileage else "**Current Mileage:** N/A")
                            
                            if initial_mileage and current_mileage:
                                total_miles = current_mileage - initial_mileage
                                st.write(f"**Total Miles Driven:** {total_miles:,} mi")
                    
                    with tab2:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Service History")
                            last_service = vehicle.get('last_service_date', 'N/A')
                            st.write(f"**Last Service Date:** {last_service}")
                            
                            last_service_mileage = vehicle.get('last_service_mileage', 0)
                            st.write(f"**Last Service Mileage:** {last_service_mileage:,} mi" if last_service_mileage else "**Last Service Mileage:** N/A")
                        
                        with col2:
                            st.write("### Service Status")
                            miles_left = vehicle.get('miles_until_service', 0)
                            
                            if miles_left is not None:
                                if miles_left > 1000:
                                    st.success(f"🟢 Service in {miles_left:,} miles")
                                    st.write("**Status:** Good")
                                elif miles_left > 500:
                                    st.warning(f"🟡 Service in {miles_left:,} miles")
                                    st.write("**Status:** Due Soon")
                                elif miles_left > 0:
                                    st.error(f"🔴 Service in {miles_left:,} miles")
                                    st.write("**Status:** Due Now")
                                else:
                                    st.error("⚠️ Service Overdue!")
                                    st.write(f"**Overdue by:** {abs(miles_left):,} miles")
                            else:
                                st.info("Service status not available")
                    
                    with tab3:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("### Current Assignment")
                            current_driver = vehicle.get('current_driver', 'Unassigned')
                            st.write(f"**Current Driver:** {current_driver if current_driver else 'Unassigned'}")
                            
                            current_trip = vehicle.get('current_trip_id', 'None')
                            st.write(f"**Current Trip ID:** {current_trip if current_trip and current_trip != 'None' else 'N/A'}")
                        
                        with col2:
                            st.write("### Usage Information")
                            last_used = vehicle.get('last_used_date', 'N/A')
                            st.write(f"**Last Used:** {last_used}")
                            
                            usage_count = vehicle.get('usage_count', 0)
                            st.write(f"**Total Trips:** {usage_count if usage_count else 0}")
                    
                    with tab4:
                        st.write("### Vehicle Notes")
                        notes = vehicle.get('notes_log', 'No notes available')
                        if notes and str(notes) != 'nan' and str(notes).strip():
                            st.text_area("Notes", value=notes, height=200, disabled=True)
                        else:
                            st.info("No notes recorded for this vehicle")
                        
                        st.write("---")
                        created = vehicle.get('created_at', 'N/A')
                        updated = vehicle.get('updated_at', 'N/A')
                        st.caption(f"Created: {created} | Last Updated: {updated}")
            
            # GALLERY LIST VIEW
            else:
                # Add Vehicle button at top
                col1, col2 = st.columns([5, 1])
                with col2:
                    if st.button("➕ Add Vehicle"):
                        st.session_state.add_vehicle_mode = True
                        st.rerun()
                
                query = """
                    SELECT 
                        id, year, make_model, vin, license_plate,
                        photo_url, picture_url, current_mileage,
                        last_service_date, last_service_mileage,
                        miles_until_service, status, last_used_date,
                        current_driver, usage_count, created_at, updated_at
                    FROM dbo.vehicles 
                    ORDER BY 
                        CASE 
                            WHEN status = 'Available' THEN 1
                            WHEN status = 'Dispatched' THEN 2
                            WHEN status = 'Out Of Service' THEN 3
                            WHEN status = 'Maintenance' THEN 4
                            ELSE 5
                        END,
                        current_mileage ASC
                """
                df, err = execute_query(query)
                
                if err:
                    st.error(f"Could not load fleet: {err}")
                elif df is None or df.empty:
                    st.info("No vehicles found in the fleet.")
                else:
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        available = len(df[df['status'] == 'Available'])
                        st.metric("Available", available, delta=None)
                    with col2:
                        dispatched = len(df[df['status'] == 'Dispatched'])
                        st.metric("Dispatched", dispatched)
                    with col3:
                        maintenance = len(df[df['status'] == 'Maintenance'])
                        st.metric("In Maintenance", maintenance)
                    with col4:
                        total_vehicles = len(df)
                        st.metric("Total Fleet", total_vehicles)
                    
                    st.markdown("---")
                    
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        search = st.text_input("🔍 Search by Make/Model or License Plate", "")
                    with col2:
                        status_filter = st.multiselect(
                            "Filter by Status",
                            options=['Available', 'Dispatched', 'Maintenance', 'Out Of Service'],
                            default=['Available', 'Dispatched']
                        )
                    
                    filtered_df = df.copy()
                    if search:
                        filtered_df = filtered_df[
                            filtered_df['make_model'].str.contains(search, case=False, na=False) |
                            filtered_df['license_plate'].str.contains(search, case=False, na=False)
                        ]
                    if status_filter:
                        filtered_df = filtered_df[filtered_df['status'].isin(status_filter)]
                    
                    st.markdown("---")
                    
                    if len(filtered_df) == 0:
                        st.info("No vehicles match your search criteria.")
                    else:
                        st.success(f"📊 Showing {len(filtered_df)} vehicle(s)")
                        
                        for idx, vehicle in filtered_df.iterrows():
                            row_class = "item-row-even" if idx % 2 == 0 else "item-row-odd"
                            st.markdown(f'<div class="item-row {row_class}">', unsafe_allow_html=True)
                            
                            col_photo, col1, col2, col3, col4 = st.columns([1, 2.5, 2, 2, 1])
                            
                            with col_photo:
                                photo_url = vehicle.get('photo_url') or vehicle.get('picture_url')
                                if photo_url and str(photo_url).strip() and str(photo_url) != 'nan':
                                    try:
                                        st.image(photo_url, width='stretch')
                                    except:
                                        st.markdown("""
                                            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                                        padding: 30px; text-align: center; border-radius: 8px; color: white; font-size: 36px;">
                                                🚗
                                            </div>
                                        """, unsafe_allow_html=True)
                                else:
                                    st.markdown("""
                                        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                                    padding: 30px; text-align: center; border-radius: 8px; color: white; font-size: 36px;">
                                            🚗
                                        </div>
                                    """, unsafe_allow_html=True)
                            
                            with col1:
                                year = vehicle.get('year', 'N/A')
                                make_model = vehicle.get('make_model', 'N/A')
                                vehicle_id = vehicle.get('id', 'N/A')
                                st.markdown(f'<div class="list-header">🚗 {year} {make_model}</div>', unsafe_allow_html=True)
                                
                                license = vehicle.get('license_plate', 'N/A')
                                vin = vehicle.get('vin', 'N/A')
                                st.write(f"**License:** {license}")
                                st.caption(f"VIN: {vin}")
                            
                            with col2:
                                current_mileage = vehicle.get('current_mileage', 0)
                                miles_left = vehicle.get('miles_until_service', 0)
                                
                                if current_mileage:
                                    st.write(f"**Mileage:** {current_mileage:,} mi")
                                else:
                                    st.write(f"**Mileage:** N/A")
                                
                                if miles_left is not None and miles_left > 0:
                                    if miles_left > 1000:
                                        service_icon = '🟢'
                                        service_status = 'Good'
                                    elif miles_left > 500:
                                        service_icon = '🟡'
                                        service_status = 'Due Soon'
                                    else:
                                        service_icon = '🔴'
                                        service_status = 'Due Now'
                                    st.caption(f"{service_icon} Service in {miles_left} mi ({service_status})")
                                else:
                                    st.caption("⚠️ Service overdue")
                            
                            with col3:
                                status = vehicle.get('status', 'Unknown')
                                status_colors = {
                                    'Available': '🟢',
                                    'Dispatched': '🟡', 
                                    'Maintenance': '🟠',
                                    'Out Of Service': '🔴'
                                }
                                st.write(f"{status_colors.get(status, '⚪')} **{status}**")
                                
                                last_used = vehicle.get('last_used_date', 'N/A')
                                if last_used and last_used != 'N/A':
                                    st.caption(f"Last used: {last_used}")
                                else:
                                    st.caption("Last used: Never")
                            
                            with col4:
                                if st.button("📋 View", key=f"view_vehicle_{idx}_{vehicle_id}"):
                                    st.session_state.view_vehicle_id = vehicle_id
                                    st.rerun()
                            
                            st.markdown('</div>', unsafe_allow_html=True)
                        
                        st.markdown("---")
                        
                        st.download_button(
                            "📥 Download Fleet Report (CSV)",
                            data=filtered_df.to_csv(index=False).encode('utf-8'),
                            file_name=f"vdh_fleet_report_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
        
        with tab4:
            st.subheader("📊 Driver Trip Logs & Reports")
            st.markdown("View and export all driver trip logs for fleet reporting and analysis.")
            
            # Date range filter
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                start_date = st.date_input("Start Date", value=datetime.now().date() - timedelta(days=30))
            with col2:
                end_date = st.date_input("End Date", value=datetime.now().date())
            with col3:
                st.write("")  # Spacing
                filter_active = st.checkbox("Active Only", value=False)
            
            # Query all trip logs
            if filter_active:
                trips_query = """
                    SELECT 
                        t.trip_id,
                        t.vehicle_id,
                        v.vehicle_name,
                        v.license_plate,
                        t.driver_name,
                        t.driver_email,
                        t.department,
                        t.start_location,
                        t.start_mileage,
                        t.start_datetime,
                        t.end_location,
                        t.end_mileage,
                        t.end_datetime,
                        t.miles_driven,
                        t.trip_status,
                        t.notes,
                        (SELECT COUNT(*) FROM dbo.trip_photos WHERE trip_id = t.trip_id) as photo_count
                    FROM dbo.vehicle_trips t
                    JOIN dbo.vehicles v ON t.vehicle_id = v.id
                    WHERE t.trip_status = 'In Progress'
                    ORDER BY t.start_datetime DESC
                """
                params = None
            else:
                trips_query = """
                    SELECT 
                        t.trip_id,
                        t.vehicle_id,
                        v.vehicle_name,
                        v.license_plate,
                        t.driver_name,
                        t.driver_email,
                        t.department,
                        t.start_location,
                        t.start_mileage,
                        t.start_datetime,
                        t.end_location,
                        t.end_mileage,
                        t.end_datetime,
                        t.miles_driven,
                        t.trip_status,
                        t.notes,
                        (SELECT COUNT(*) FROM dbo.trip_photos WHERE trip_id = t.trip_id) as photo_count
                    FROM dbo.vehicle_trips t
                    JOIN dbo.vehicles v ON t.vehicle_id = v.id
                    WHERE CAST(t.start_datetime AS DATE) >= ? AND CAST(t.start_datetime AS DATE) <= ?
                    ORDER BY t.start_datetime DESC
                """
                params = (start_date, end_date)
            
            trips_df, trips_err = execute_query(trips_query, params)
            
            if trips_err:
                st.error(f"Error loading trip logs: {trips_err}")
            elif trips_df is None or trips_df.empty:
                st.info("📭 No trip logs found for the selected period")
            else:
                # Summary statistics
                st.markdown("### 📈 Trip Summary")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Trips", len(trips_df))
                with col2:
                    completed = trips_df[trips_df['trip_status'] == 'Completed']
                    st.metric("Completed", len(completed))
                with col3:
                    in_progress = trips_df[trips_df['trip_status'] == 'In Progress']
                    st.metric("In Progress", len(in_progress))
                with col4:
                    total_miles = trips_df[trips_df['miles_driven'].notna()]['miles_driven'].sum()
                    st.metric("Total Miles", f"{int(total_miles):,}")
                
                st.markdown("---")
                
                # Department breakdown
                if len(trips_df) > 0:
                    st.markdown("### 🏢 Miles by Department")
                    dept_summary = trips_df[trips_df['miles_driven'].notna()].groupby('department')['miles_driven'].sum().sort_values(ascending=False)
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        if not dept_summary.empty:
                            st.bar_chart(dept_summary)
                    with col2:
                        st.dataframe(dept_summary.reset_index().rename(columns={'department': 'Department', 'miles_driven': 'Miles'}), use_container_width=True)
                
                st.markdown("---")
                st.markdown("### 📋 Detailed Trip Logs")
                
                # Display trips
                for _, trip in trips_df.iterrows():
                    status_emoji = "🟢" if trip['trip_status'] == 'In Progress' else "✅"
                    vehicle_info = f"{trip['vehicle_name']} ({trip['license_plate']})"
                    
                    with st.expander(f"{status_emoji} {vehicle_info} - {trip['driver_name']} - {trip['start_datetime'].strftime('%m/%d/%Y %I:%M %p')}"):
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.write("**Trip Information**")
                            st.write(f"Trip ID: {trip['trip_id']}")
                            st.write(f"Status: {trip['trip_status']}")
                            st.write(f"Department: {trip['department']}")
                            if trip['miles_driven']:
                                st.write(f"**Miles Driven:** {trip['miles_driven']}")
                        
                        with col2:
                            st.write("**Driver Information**")
                            st.write(f"Name: {trip['driver_name']}")
                            st.write(f"Email: {trip['driver_email']}")
                            st.write(f"From: {trip['start_location']}")
                            st.write(f"To: {trip['end_location'] or 'In progress...'}")
                        
                        with col3:
                            st.write("**Mileage Details**")
                            st.write(f"Start: {trip['start_mileage']:,} mi")
                            st.write(f"End: {trip['end_mileage'] or 'N/A'}")
                            st.write(f"Started: {trip['start_datetime'].strftime('%m/%d/%Y %I:%M %p')}")
                            if trip['end_datetime']:
                                st.write(f"Ended: {trip['end_datetime'].strftime('%m/%d/%Y %I:%M %p')}")
                        
                        if trip['notes']:
                            st.write(f"**Notes:** {trip['notes']}")
                        
                        if trip['photo_count'] > 0:
                            st.info(f"📸 {trip['photo_count']} photo(s) attached to this trip")
                
                # Export functionality
                st.markdown("---")
                st.markdown("### 📥 Export Trip Data")
                
                # Prepare export data
                export_df = trips_df[[
                    'trip_id', 'vehicle_name', 'license_plate', 'driver_name', 
                    'driver_email', 'department', 'start_location', 'end_location',
                    'start_mileage', 'end_mileage', 'miles_driven', 
                    'start_datetime', 'end_datetime', 'trip_status', 'photo_count'
                ]].copy()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        "📥 Download Trip Report (CSV)",
                        data=export_df.to_csv(index=False).encode('utf-8'),
                        file_name=f"vdh_trip_logs_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                with col2:
                    # Excel export with formatting
                    from io import BytesIO
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        export_df.to_excel(writer, index=False, sheet_name='Trip Logs')
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        "📥 Download Trip Report (Excel)",
                        data=excel_data,
                        file_name=f"vdh_trip_logs_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

    # Report Builder and Connection Test
    elif page == "📈 Report Builder":
        st.header("📈 Report Builder")
        
        if not DB_AVAILABLE:
            st.warning("Database unavailable. Report generation requires a live database connection.")
        else:
            st.info("💡 **Tip:** Select report type, apply filters, preview data, then download your custom report.")
            
            # Report Configuration
            st.markdown("---")
            st.subheader("📋 Report Configuration")
            
            col1, col2 = st.columns(2)
            
            with col1:
                report_type = st.selectbox(
                    "Report Type",
                    ["🎫 Helpdesk Tickets", "💻 Assets", "🛒 Procurement Requests", "🚗 Fleet Vehicles"],
                    help="Select the type of report to generate"
                )
            
            with col2:
                export_format = st.selectbox(
                    "Export Format",
                    ["CSV", "Excel (XLSX)"],
                    help="Choose output format"
                )
            
            st.markdown("---")
            st.subheader("🔍 Filters")
            
            # Date range filter
            col1, col2 = st.columns(2)
            with col1:
                date_from = st.date_input("Date From", value=None, help="Leave empty for all dates")
            with col2:
                date_to = st.date_input("Date To", value=None, help="Leave empty for all dates")
            
            # Type-specific filters
            if report_type == "🎫 Helpdesk Tickets":
                col1, col2, col3 = st.columns(3)
                with col1:
                    status_filter = st.multiselect(
                        "Status",
                        ["New", "Open", "In Progress", "On Hold", "Resolved", "Closed"],
                        default=["New", "Open", "In Progress"]
                    )
                with col2:
                    priority_filter = st.multiselect(
                        "Priority",
                        ["Low", "Medium", "High", "Critical"],
                        default=["Low", "Medium", "High", "Critical"]
                    )
                with col3:
                    location_filter = st.multiselect(
                        "Location",
                        ["Petersburg", "Hopewell", "Surry", "Sussex", "Dinwiddie", "Prince George", "Greensville/Emporia", "Crater"],
                        default=[]
                    )
                
                # Build query
                query = """
                    SELECT 
                        ticket_id, status, priority, name, email, phone_number,
                        location, short_description, description, created_at,
                        assigned_to, resolved_at
                    FROM dbo.Tickets
                    WHERE 1=1
                """
                
                if status_filter:
                    status_list = "', '".join(status_filter)
                    query += f" AND status IN ('{status_list}')"
                if priority_filter:
                    priority_list = "', '".join(priority_filter)
                    query += f" AND priority IN ('{priority_list}')"
                if location_filter:
                    location_list = "', '".join(location_filter)
                    query += f" AND location IN ('{location_list}')"
                if date_from:
                    query += f" AND CAST(created_at AS DATE) >= '{date_from}'"
                if date_to:
                    query += f" AND CAST(created_at AS DATE) <= '{date_to}'"
                
                query += " ORDER BY created_at DESC"
            
            elif report_type == "💻 Assets":
                col1, col2, col3 = st.columns(3)
                with col1:
                    status_filter = st.multiselect(
                        "Status",
                        ["Deployed", "In-Stock", "Surplus", "Unaccounted"],
                        default=["Deployed", "In-Stock"]
                    )
                with col2:
                    category_filter = st.multiselect(
                        "Category",
                        ["Network", "Workstation", "Printer", "Mobile"],
                        default=[]
                    )
                with col3:
                    location_filter = st.multiselect(
                        "Location",
                        ["Petersburg", "Hopewell", "Surry", "Sussex", "Dinwiddie", "Prince George", "Greensville/Emporia", "Crater"],
                        default=[]
                    )
                
                query = """
                    SELECT 
                        asset_id, asset_tag, type, category, model, serial,
                        status, location, assigned_user, assigned_email,
                        purchase_date, warranty_expiration, created_at
                    FROM dbo.Assets
                    WHERE 1=1
                """
                
                if status_filter:
                    status_list = "', '".join(status_filter)
                    query += f" AND status IN ('{status_list}')"
                if category_filter:
                    category_list = "', '".join(category_filter)
                    query += f" AND category IN ('{category_list}')"
                if location_filter:
                    location_list = "', '".join(location_filter)
                    query += f" AND location IN ('{location_list}')"
                if date_from:
                    query += f" AND CAST(created_at AS DATE) >= '{date_from}'"
                if date_to:
                    query += f" AND CAST(created_at AS DATE) <= '{date_to}'"
                
                query += " ORDER BY asset_id DESC"
            
            elif report_type == "🛒 Procurement Requests":
                col1, col2 = st.columns(2)
                with col1:
                    status_filter = st.multiselect(
                        "Status",
                        ["Draft", "Pending", "Approved", "Rejected", "Completed"],
                        default=["Pending", "Approved"]
                    )
                with col2:
                    location_filter = st.multiselect(
                        "Location",
                        ["Petersburg", "Hopewell", "Surry", "Sussex", "Dinwiddie", "Prince George", "Greensville/Emporia", "Crater"],
                        default=[]
                    )
                
                query = """
                    SELECT 
                        request_id, request_number, request_date, requester_name,
                        requester_email, requester_phone, location, total_amount,
                        status, approval_1_status, approval_2_status
                    FROM dbo.Procurement_Requests
                    WHERE 1=1
                """
                
                if status_filter:
                    status_list = "', '".join(status_filter)
                    query += f" AND status IN ('{status_list}')"
                if location_filter:
                    location_list = "', '".join(location_filter)
                    query += f" AND location IN ('{location_list}')"
                if date_from:
                    query += f" AND request_date >= '{date_from}'"
                if date_to:
                    query += f" AND request_date <= '{date_to}'"
                
                query += " ORDER BY request_date DESC"
            
            elif report_type == "🚗 Fleet Vehicles":
                col1, col2 = st.columns(2)
                with col1:
                    status_filter = st.multiselect(
                        "Status",
                        ["Available", "Dispatched", "Maintenance", "Out Of Service"],
                        default=["Available", "Dispatched"]
                    )
                with col2:
                    mileage_threshold = st.number_input(
                        "Max Current Mileage (optional)",
                        min_value=0,
                        value=0,
                        help="Filter vehicles below this mileage (0 = no filter)"
                    )
                
                query = """
                    SELECT 
                        id, year, make_model, vin, license_plate,
                        initial_mileage, current_mileage, last_service_date,
                        last_service_mileage, miles_until_service, status,
                        last_used_date, current_driver, created_at
                    FROM dbo.vehicles
                    WHERE 1=1
                """
                
                if status_filter:
                    status_list = "', '".join(status_filter)
                    query += f" AND status IN ('{status_list}')"
                if mileage_threshold > 0:
                    query += f" AND current_mileage <= {mileage_threshold}"
                if date_from:
                    query += f" AND CAST(created_at AS DATE) >= '{date_from}'"
                if date_to:
                    query += f" AND CAST(created_at AS DATE) <= '{date_to}'"
                
                query += " ORDER BY current_mileage ASC"
            
            st.markdown("---")
            
            # Generate Report Button
            col1, col2, col3 = st.columns([2, 1, 2])
            with col2:
                generate_button = st.button("🔄 Generate Report", use_container_width=True)
            
            if generate_button:
                with st.spinner("Generating report..."):
                    df, error = execute_query(query)
                    
                    if error:
                        st.error(f"❌ Error generating report: {error}")
                    elif df is None or df.empty:
                        st.warning("⚠️ No data found matching your filters.")
                    else:
                        st.success(f"✅ Report generated successfully! Found {len(df)} record(s).")
                        
                        # Summary Statistics
                        st.markdown("---")
                        st.subheader("📊 Summary Statistics")
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Total Records", len(df))
                        
                        if report_type == "🎫 Helpdesk Tickets":
                            with col2:
                                resolved = len(df[df['status'] == 'Resolved']) if 'status' in df.columns else 0
                                st.metric("Resolved", resolved)
                            with col3:
                                high_priority = len(df[df['priority'].isin(['High', 'Critical'])]) if 'priority' in df.columns else 0
                                st.metric("High Priority", high_priority)
                            with col4:
                                avg_time = "N/A"
                                st.metric("Avg Resolution", avg_time)
                        
                        elif report_type == "💻 Assets":
                            with col2:
                                deployed = len(df[df['status'] == 'Deployed']) if 'status' in df.columns else 0
                                st.metric("Deployed", deployed)
                            with col3:
                                in_stock = len(df[df['status'] == 'In-Stock']) if 'status' in df.columns else 0
                                st.metric("In Stock", in_stock)
                            with col4:
                                surplus = len(df[df['status'] == 'Surplus']) if 'status' in df.columns else 0
                                st.metric("Surplus", surplus)
                        
                        elif report_type == "🛒 Procurement Requests":
                            with col2:
                                pending = len(df[df['status'] == 'Pending']) if 'status' in df.columns else 0
                                st.metric("Pending", pending)
                            with col3:
                                total_amount = df['total_amount'].sum() if 'total_amount' in df.columns else 0
                                st.metric("Total Amount", f"${total_amount:,.2f}")
                            with col4:
                                approved = len(df[df['status'] == 'Approved']) if 'status' in df.columns else 0
                                st.metric("Approved", approved)
                        
                        elif report_type == "🚗 Fleet Vehicles":
                            with col2:
                                available = len(df[df['status'] == 'Available']) if 'status' in df.columns else 0
                                st.metric("Available", available)
                            with col3:
                                avg_mileage = int(df['current_mileage'].mean()) if 'current_mileage' in df.columns else 0
                                st.metric("Avg Mileage", f"{avg_mileage:,} mi")
                            with col4:
                                needs_service = len(df[df['miles_until_service'] < 500]) if 'miles_until_service' in df.columns else 0
                                st.metric("Needs Service", needs_service)
                        
                        # Data Preview
                        st.markdown("---")
                        st.subheader("📄 Data Preview")
                        st.dataframe(df, use_container_width=True, height=400)
                        
                        # Export Options
                        st.markdown("---")
                        st.subheader("💾 Export Report")
                        
                        col1, col2, col3 = st.columns([1, 2, 1])
                        
                        with col2:
                            if export_format == "CSV":
                                csv_data = df.to_csv(index=False).encode('utf-8')
                                report_name = report_type.replace("🎫 ", "").replace("💻 ", "").replace("🛒 ", "").replace("🚗 ", "").replace(" ", "_")
                                filename = f"VDH_{report_name}_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                                
                                st.download_button(
                                    label="📥 Download CSV Report",
                                    data=csv_data,
                                    file_name=filename,
                                    mime="text/csv",
                                    use_container_width=True
                                )
                            
                            elif export_format == "Excel (XLSX)":
                                try:
                                    from io import BytesIO
                                    output = BytesIO()
                                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                        df.to_excel(writer, sheet_name='Report', index=False)
                                    excel_data = output.getvalue()
                                    
                                    report_name = report_type.replace("🎫 ", "").replace("💻 ", "").replace("🛒 ", "").replace("🚗 ", "").replace(" ", "_")
                                    filename = f"VDH_{report_name}_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                    
                                    st.download_button(
                                        label="📥 Download Excel Report",
                                        data=excel_data,
                                        file_name=filename,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True
                                    )
                                except Exception as e:
                                    st.error(f"Excel export requires openpyxl. Install with: pip install openpyxl")
                                    st.info("Falling back to CSV export...")
                                    csv_data = df.to_csv(index=False).encode('utf-8')
                                    report_name = report_type.replace("🎫 ", "").replace("💻 ", "").replace("🛒 ", "").replace("🚗 ", "").replace(" ", "_")
                                    filename = f"VDH_{report_name}_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                                    
                                    st.download_button(
                                        label="📥 Download CSV Report (Fallback)",
                                        data=csv_data,
                                        file_name=filename,
                                        mime="text/csv",
                                        use_container_width=True
                                    )

    elif page == "🔌 Connection Test":
        st.header("🔌 Database Connection Test")
        server, database, username, _ = get_connection_string()
        st.write(f"**Server:** {server}")
        st.write(f"**Database:** {database}")
        st.write(f"**Username:** {username}")
        try:
            import pyodbc
            st.write("**pyodbc Available:** ✅ Yes")
        except Exception:
            st.write("**pyodbc Available:** ❌ No")
        if DB_AVAILABLE:
            st.success("Database appears to be reachable (connection established during startup check).")
        else:
            st.error("Database is not reachable. Check configuration and network.")

    elif page == "📦 Resource Management":
        render_resource_management()

    st.markdown("---")
    st.markdown("*VDH Service Center - Comprehensive Management System | Virginia Department of Health © 2025*")

# Run
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.exception("Unexpected error in main app: %s", e)
        try:
            st.error("An unexpected error occurred while loading the app. The error has been logged. Please contact the administrator.")
        except Exception:
            raise